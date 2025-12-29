from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.future import select
from sqlalchemy.orm import selectinload
from sqlalchemy import and_, extract, or_, cast, Date
from typing import Optional, List
from datetime import datetime, date, timedelta
from io import BytesIO
import models
from database import get_db

# Try to import openpyxl for Excel generation
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

router = APIRouter()

# Spanish month names
MESES = {
    1: 'Enero', 2: 'Febrero', 3: 'Marzo', 4: 'Abril',
    5: 'Mayo', 6: 'Junio', 7: 'Julio', 8: 'Agosto',
    9: 'Septiembre', 10: 'Octubre', 11: 'Noviembre', 12: 'Diciembre'
}


async def get_report_data(
    db: AsyncSession,
    proveedor_id: Optional[int] = None,
    oficina_id: Optional[int] = None,
    fecha_desde: Optional[date] = None,
    fecha_hasta: Optional[date] = None,
    año: Optional[int] = None,
    mes: Optional[int] = None,
    tipo: Optional[str] = None,
    estado: Optional[str] = None,
    ciudad: Optional[str] = None
):
    """
    Get report data with all contracts and their invoice values from FacturaOficina.
    Returns data structured for Excel export with dynamic month columns.
    """
    # Determine date range for columns FIRST
    if fecha_desde and fecha_hasta:
        start_date = fecha_desde
        end_date = fecha_hasta
    elif año and mes:
        # Specific month
        start_date = date(año, mes, 1)
        # Get last day of month
        if mes == 12:
            end_date = date(año, 12, 31)
        else:
            end_date = date(año, mes + 1, 1) - timedelta(days=1)
    elif año:
        start_date = date(año, 1, 1)
        end_date = date(año, 12, 31)
    else:
        # Default: current year
        current_year = datetime.now().year
        start_date = date(current_year, 1, 1)
        end_date = date(current_year, 12, 31)
    
    # Generate list of months in date range
    months = []
    current = date(start_date.year, start_date.month, 1)
    while current <= end_date:
        months.append((current.year, current.month))
        # Move to next month
        if current.month == 12:
            current = date(current.year + 1, 1, 1)
        else:
            current = date(current.year, current.month + 1, 1)
    
    # Build query for contracts with relationships
    query = (
        select(models.Contrato)
        .options(
            selectinload(models.Contrato.proveedor),
            selectinload(models.Contrato.oficina)
        )
        .outerjoin(models.Oficina)
    )
    
    # Apply filters
    filters = []
    
    if proveedor_id:
        filters.append(models.Contrato.proveedor_id == proveedor_id)
    
    if oficina_id:
        filters.append(models.Contrato.oficina_id == oficina_id)
    
    if tipo:
        filters.append(models.Contrato.tipo == tipo)
    
    if estado:
        filters.append(models.Contrato.estado == estado)
    
    if ciudad:
        filters.append(models.Oficina.ciudad.ilike(f"%{ciudad}%"))
    
    if filters:
        query = query.filter(and_(*filters))
    
    result = await db.execute(query.order_by(models.Contrato.id))
    contratos = result.scalars().all()
    
    # Get all FacturaOficina entries with their facturas for the date range
    # This gets all invoice assignments to offices within the date range
    factura_oficina_query = (
        select(models.FacturaOficina)
        .options(
            selectinload(models.FacturaOficina.factura),
            selectinload(models.FacturaOficina.oficina),
            selectinload(models.FacturaOficina.contrato)
        )
        .join(models.Factura)
        .outerjoin(models.Oficina, models.FacturaOficina.oficina_id == models.Oficina.id)
    )
    
    # Apply date filter - use fecha_factura or created_at if fecha_factura is NULL
    # Get the effective date (fecha_factura or created_at date)
    fo_filters = [
        or_(
            and_(
                models.Factura.fecha_factura.isnot(None),
                models.Factura.fecha_factura >= start_date,
                models.Factura.fecha_factura <= end_date
            ),
            and_(
                models.Factura.fecha_factura.is_(None),
                cast(models.Factura.created_at, Date) >= start_date,
                cast(models.Factura.created_at, Date) <= end_date
            )
        )
    ]
    
    # Apply same filters to FacturaOficina query
    if proveedor_id:
        fo_filters.append(models.Factura.proveedor_id == proveedor_id)
    
    if oficina_id:
        fo_filters.append(models.FacturaOficina.oficina_id == oficina_id)
    
    if ciudad:
        fo_filters.append(models.Oficina.ciudad.ilike(f"%{ciudad}%"))
    
    factura_oficina_query = factura_oficina_query.filter(and_(*fo_filters))
    
    fo_result = await db.execute(factura_oficina_query)
    factura_oficinas = fo_result.scalars().all()
    
    # Create a lookup: (proveedor_id, oficina_id) -> {month_key: {valor, fecha}}
    # This groups all invoice values by proveedor+oficina+month
    valores_por_contrato = {}
    for fo in factura_oficinas:
        if fo.factura:
            # Use fecha_factura if available, otherwise use created_at
            if fo.factura.fecha_factura:
                factura_fecha = fo.factura.fecha_factura
            elif fo.factura.created_at:
                factura_fecha = fo.factura.created_at.date() if hasattr(fo.factura.created_at, 'date') else fo.factura.created_at
            else:
                continue  # Skip if no date available
            
            proveedor_id_fo = fo.factura.proveedor_id
            oficina_id_fo = fo.oficina_id
            
            key = (proveedor_id_fo, oficina_id_fo)
            month_key = f"{factura_fecha.year}-{factura_fecha.month:02d}"
            
            if key not in valores_por_contrato:
                valores_por_contrato[key] = {}
            
            if month_key not in valores_por_contrato[key]:
                valores_por_contrato[key][month_key] = {'valor': 0, 'fecha': None, 'facturas': []}
            
            valor_fo = float(fo.valor) if fo.valor else 0
            valores_por_contrato[key][month_key]['valor'] += valor_fo
            valores_por_contrato[key][month_key]['fecha'] = factura_fecha.isoformat() if hasattr(factura_fecha, 'isoformat') else str(factura_fecha)
            if fo.factura.numero_factura:
                valores_por_contrato[key][month_key]['facturas'].append(fo.factura.numero_factura)
    
    # Process data - now using the lookup
    report_data = []
    for contrato in contratos:
        row = {
            'nit_proveedor': contrato.proveedor.nit if contrato.proveedor else '',
            'nombre_proveedor': contrato.proveedor.nombre if contrato.proveedor else '',
            'cod_oficina': contrato.oficina.cod_oficina if contrato.oficina else '',
            'nombre_oficina': contrato.oficina.nombre if contrato.oficina else '',
            'direccion': contrato.oficina.direccion if contrato.oficina else '',
            'ciudad': contrato.oficina.ciudad if contrato.oficina else '',
            'tipo': contrato.tipo or '',
            'num_contrato': contrato.num_contrato or '',
            'tipo_plan': contrato.tipo_plan or '',
            'tipo_canal': contrato.tipo_canal or '',
            'valor_mensual': float(contrato.valor_mensual) if contrato.valor_mensual else 0,
            'pagos': {}
        }
        
        # Get invoice values for this contrato's proveedor+oficina combination
        key = (contrato.proveedor_id, contrato.oficina_id)
        if key in valores_por_contrato:
            row['pagos'] = {
                mk: {'valor': v['valor'], 'fecha': v['fecha']} 
                for mk, v in valores_por_contrato[key].items()
            }
        

        report_data.append(row)
    
    return report_data, months


def create_excel_report(data: List[dict], months: List[tuple], titulo: str = "Reporte de Contratos"):
    """Create Excel workbook with report data"""
    if not EXCEL_AVAILABLE:
        raise ImportError("openpyxl is not installed. Run: pip install openpyxl")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    month_header_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
    date_header_fill = PatternFill(start_color="F59E0B", end_color="F59E0B", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    currency_font = Font(bold=True)
    
    # Static headers
    static_headers = [
        'NIT Proveedor',
        'Nombre Proveedor',
        'Código Oficina',
        'Nombre Oficina',
        'Dirección',
        'Ciudad',
        'Tipo',
        'Número Contrato',
        'Tipo Plan',
        'Tipo Canal',
        'Valor Mensual'
    ]
    
    # Write headers
    col = 1
    for header in static_headers:
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        col += 1
    
    # Dynamic month headers
    for year, month in months:
        month_name = MESES[month]
        year_short = str(year)[-2:]  # Last 2 digits of year
        
        # Valor column
        cell = ws.cell(row=1, column=col, value=f"Valor {month_name} {year_short}")
        cell.font = header_font
        cell.fill = month_header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        col += 1
        
        # Fecha column
        cell = ws.cell(row=1, column=col, value=f"Fecha {month_name} {year_short}")
        cell.font = header_font
        cell.fill = date_header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        col += 1
    
    # Write data rows
    row_num = 2
    for item in data:
        col = 1
        
        # Static columns
        ws.cell(row=row_num, column=col, value=item['nit_proveedor']).border = thin_border; col += 1
        ws.cell(row=row_num, column=col, value=item['nombre_proveedor']).border = thin_border; col += 1
        ws.cell(row=row_num, column=col, value=item['cod_oficina']).border = thin_border; col += 1
        ws.cell(row=row_num, column=col, value=item['nombre_oficina']).border = thin_border; col += 1
        ws.cell(row=row_num, column=col, value=item['direccion']).border = thin_border; col += 1
        ws.cell(row=row_num, column=col, value=item['ciudad']).border = thin_border; col += 1
        ws.cell(row=row_num, column=col, value=item['tipo']).border = thin_border; col += 1
        ws.cell(row=row_num, column=col, value=item['num_contrato']).border = thin_border; col += 1
        ws.cell(row=row_num, column=col, value=item['tipo_plan']).border = thin_border; col += 1
        ws.cell(row=row_num, column=col, value=item['tipo_canal']).border = thin_border; col += 1
        
        valor_cell = ws.cell(row=row_num, column=col, value=item['valor_mensual'])
        valor_cell.border = thin_border
        valor_cell.number_format = '#,##0'
        col += 1
        
        # Dynamic month columns
        for year, month in months:
            # Use same string key format as data
            month_key = f"{year}-{month:02d}"
            pago_data = item['pagos'].get(month_key, {})
            
            # Valor
            valor = pago_data.get('valor', 0)
            valor_cell = ws.cell(row=row_num, column=col, value=valor if valor > 0 else '')
            valor_cell.border = thin_border
            if valor > 0:
                valor_cell.number_format = '#,##0'
                valor_cell.font = currency_font
            col += 1
            
            # Fecha
            fecha = pago_data.get('fecha')
            fecha_cell = ws.cell(row=row_num, column=col, value=fecha if fecha else '')
            fecha_cell.border = thin_border
            col += 1
        
        row_num += 1
    
    # Adjust column widths
    for col_idx, header in enumerate(static_headers, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(len(header) + 2, 15)
    
    # Adjust month columns
    month_col_start = len(static_headers) + 1
    for i, (year, month) in enumerate(months):
        valor_col = month_col_start + (i * 2)
        fecha_col = valor_col + 1
        ws.column_dimensions[get_column_letter(valor_col)].width = 15
        ws.column_dimensions[get_column_letter(fecha_col)].width = 15
    
    # Freeze first row
    ws.freeze_panes = 'A2'
    
    return wb


@router.get("/reportes/preview")
async def preview_report(
    proveedor_id: Optional[int] = Query(None, description="Filtrar por proveedor"),
    oficina_id: Optional[int] = Query(None, description="Filtrar por oficina"),
    fecha_desde: Optional[str] = Query(None, description="Fecha inicio (YYYY-MM-DD)"),
    fecha_hasta: Optional[str] = Query(None, description="Fecha fin (YYYY-MM-DD)"),
    año: Optional[int] = Query(None, description="Año del reporte"),
    mes: Optional[int] = Query(None, description="Mes del reporte (1-12)"),
    tipo: Optional[str] = Query(None, description="Tipo de contrato (Fijo, Movil, Colaboracion, Leasing)"),
    estado: Optional[str] = Query(None, description="Estado del contrato (ACTIVO, CANCELADO)"),
    ciudad: Optional[str] = Query(None, description="Ciudad de la oficina"),
    db: AsyncSession = Depends(get_db)
):
    """Preview report data as JSON"""
    # Parse dates
    fecha_desde_dt = datetime.strptime(fecha_desde, '%Y-%m-%d').date() if fecha_desde else None
    fecha_hasta_dt = datetime.strptime(fecha_hasta, '%Y-%m-%d').date() if fecha_hasta else None
    
    data, months = await get_report_data(
        db,
        proveedor_id=proveedor_id,
        oficina_id=oficina_id,
        fecha_desde=fecha_desde_dt,
        fecha_hasta=fecha_hasta_dt,
        año=año,
        mes=mes,
        tipo=tipo,
        estado=estado,
        ciudad=ciudad
    )
    
    # Format months for response
    months_formatted = [
        {'year': y, 'month': m, 'name': f"{MESES[m]} {y}"} 
        for y, m in months
    ]
    
    return {
        "total_registros": len(data),
        "meses": months_formatted,
        "data": data[:50]  # Return first 50 for preview
    }


@router.get("/reportes/export")
async def export_report_excel(
    proveedor_id: Optional[int] = Query(None, description="Filtrar por proveedor"),
    oficina_id: Optional[int] = Query(None, description="Filtrar por oficina"),
    fecha_desde: Optional[str] = Query(None, description="Fecha inicio (YYYY-MM-DD)"),
    fecha_hasta: Optional[str] = Query(None, description="Fecha fin (YYYY-MM-DD)"),
    año: Optional[int] = Query(None, description="Año del reporte"),
    mes: Optional[int] = Query(None, description="Mes del reporte (1-12)"),
    tipo: Optional[str] = Query(None, description="Tipo de contrato (Fijo, Movil, Colaboracion, Leasing)"),
    estado: Optional[str] = Query(None, description="Estado del contrato (ACTIVO, CANCELADO)"),
    ciudad: Optional[str] = Query(None, description="Ciudad de la oficina"),
    db: AsyncSession = Depends(get_db)
):
    """Export report to Excel file"""
    if not EXCEL_AVAILABLE:
        return {"error": "openpyxl not installed. Run: pip install openpyxl"}
    
    # Parse dates
    fecha_desde_dt = datetime.strptime(fecha_desde, '%Y-%m-%d').date() if fecha_desde else None
    fecha_hasta_dt = datetime.strptime(fecha_hasta, '%Y-%m-%d').date() if fecha_hasta else None
    
    data, months = await get_report_data(
        db,
        proveedor_id=proveedor_id,
        oficina_id=oficina_id,
        fecha_desde=fecha_desde_dt,
        fecha_hasta=fecha_hasta_dt,
        año=año,
        mes=mes,
        tipo=tipo,
        estado=estado,
        ciudad=ciudad
    )
    
    # Create Excel
    wb = create_excel_report(data, months)
    
    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # Generate filename
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"reporte_contratos_{timestamp}.xlsx"
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/reportes/filtros")
async def get_report_filters(db: AsyncSession = Depends(get_db)):
    """Get available filter options for reports"""
    # Get all providers
    proveedores_result = await db.execute(
        select(models.Proveedor).order_by(models.Proveedor.nombre)
    )
    proveedores = proveedores_result.scalars().all()
    
    # Get all offices
    oficinas_result = await db.execute(
        select(models.Oficina).order_by(models.Oficina.nombre)
    )
    oficinas = oficinas_result.scalars().all()
    
    # Get available years from payments
    años_result = await db.execute(
        select(extract('year', models.Pago.fecha_pago).label('year'))
        .distinct()
        .order_by(extract('year', models.Pago.fecha_pago).desc())
    )
    años = [int(row[0]) for row in años_result.all() if row[0]]
    
    # Add current year if not in list
    current_year = datetime.now().year
    if current_year not in años:
        años.insert(0, current_year)
    
    # Get unique cities
    ciudades_result = await db.execute(
        select(models.Oficina.ciudad)
        .distinct()
        .filter(models.Oficina.ciudad.isnot(None))
        .order_by(models.Oficina.ciudad)
    )
    ciudades = [row[0] for row in ciudades_result.all() if row[0]]
    
    # Contract types and states
    tipos = ['Fijo', 'Movil', 'Colaboracion', 'Leasing']
    estados = ['ACTIVO', 'CANCELADO']
    
    # Months
    meses = [
        {'value': i, 'label': MESES[i]} for i in range(1, 13)
    ]
    
    return {
        "proveedores": [{"id": p.id, "nit": p.nit, "nombre": p.nombre} for p in proveedores],
        "oficinas": [{"id": o.id, "cod_oficina": o.cod_oficina, "nombre": o.nombre, "ciudad": o.ciudad} for o in oficinas],
        "años": años,
        "meses": meses,
        "tipos": tipos,
        "estados": estados,
        "ciudades": ciudades
    }


@router.get("/reportes/estadisticas")
async def get_report_stats(
    año: Optional[int] = Query(None, description="Año para estadísticas"),
    oficina_id: Optional[int] = Query(None, description="Filtrar por oficina"),
    proveedor_id: Optional[int] = Query(None, description="Filtrar por proveedor"),
    db: AsyncSession = Depends(get_db)
):
    """Get dashboard statistics for reports"""
    from sqlalchemy import func
    
    # Use current year if not specified
    target_year = año or datetime.now().year
    start_date = date(target_year, 1, 1)
    end_date = date(target_year, 12, 31)
    
    # Base date filter
    def get_date_filter():
        return or_(
            and_(
                models.Factura.fecha_factura.isnot(None),
                models.Factura.fecha_factura >= start_date,
                models.Factura.fecha_factura <= end_date
            ),
            and_(
                models.Factura.fecha_factura.is_(None),
                cast(models.Factura.created_at, Date) >= start_date,
                cast(models.Factura.created_at, Date) <= end_date
            )
        )
    
    # Oficina filter
    def get_oficina_filter():
        if oficina_id:
            return models.FacturaOficina.oficina_id == oficina_id
        return True  # No filter
    
    # 1. Total facturado en el año (from FacturaOficina)
    total_facturado_result = await db.execute(
        select(func.sum(models.FacturaOficina.valor))
        .join(models.Factura)
        .filter(
            or_(
                and_(
                    models.Factura.fecha_factura.isnot(None),
                    models.Factura.fecha_factura >= start_date,
                    models.Factura.fecha_factura <= end_date
                ),
                and_(
                    models.Factura.fecha_factura.is_(None),
                    cast(models.Factura.created_at, Date) >= start_date,
                    cast(models.Factura.created_at, Date) <= end_date
                )
            )
        )
    )
    total_facturado = float(total_facturado_result.scalar() or 0)
    
    # 2. Total de facturas en el año
    total_facturas_result = await db.execute(
        select(func.count(func.distinct(models.Factura.id)))
        .join(models.FacturaOficina)
        .filter(
            or_(
                and_(
                    models.Factura.fecha_factura.isnot(None),
                    models.Factura.fecha_factura >= start_date,
                    models.Factura.fecha_factura <= end_date
                ),
                and_(
                    models.Factura.fecha_factura.is_(None),
                    cast(models.Factura.created_at, Date) >= start_date,
                    cast(models.Factura.created_at, Date) <= end_date
                )
            )
        )
    )
    total_facturas = total_facturas_result.scalar() or 0
    
    # 3. Proveedores únicos facturados en el año
    proveedores_facturados_result = await db.execute(
        select(func.count(func.distinct(models.Factura.proveedor_id)))
        .join(models.FacturaOficina)
        .filter(
            or_(
                and_(
                    models.Factura.fecha_factura.isnot(None),
                    models.Factura.fecha_factura >= start_date,
                    models.Factura.fecha_factura <= end_date
                ),
                and_(
                    models.Factura.fecha_factura.is_(None),
                    cast(models.Factura.created_at, Date) >= start_date,
                    cast(models.Factura.created_at, Date) <= end_date
                )
            )
        )
    )
    proveedores_facturados = proveedores_facturados_result.scalar() or 0
    
    # 4. Contratos activos sin facturas en el período
    contratos_activos_result = await db.execute(
        select(func.count(models.Contrato.id))
        .filter(models.Contrato.estado == 'ACTIVO')
    )
    contratos_activos = contratos_activos_result.scalar() or 0
    
    # 5. Facturación por mes del año
    facturacion_por_mes = []
    for mes in range(1, 13):
        mes_start = date(target_year, mes, 1)
        if mes == 12:
            mes_end = date(target_year, 12, 31)
        else:
            mes_end = date(target_year, mes + 1, 1) - timedelta(days=1)
        
        # Build query with optional oficina filter
        mes_query = (
            select(func.sum(models.FacturaOficina.valor))
            .join(models.Factura)
            .filter(
                or_(
                    and_(
                        models.Factura.fecha_factura.isnot(None),
                        models.Factura.fecha_factura >= mes_start,
                        models.Factura.fecha_factura <= mes_end
                    ),
                    and_(
                        models.Factura.fecha_factura.is_(None),
                        cast(models.Factura.created_at, Date) >= mes_start,
                        cast(models.Factura.created_at, Date) <= mes_end
                    )
                )
            )
        )
        
        # Add oficina filter if specified
        if oficina_id:
            mes_query = mes_query.filter(models.FacturaOficina.oficina_id == oficina_id)
        
        # Add proveedor filter if specified
        if proveedor_id:
            mes_query = mes_query.filter(models.Factura.proveedor_id == proveedor_id)
        
        mes_result = await db.execute(mes_query)
        valor = float(mes_result.scalar() or 0)
        facturacion_por_mes.append({
            'mes': mes,
            'nombre': MESES[mes],
            'valor': valor
        })
    
    # 6. Top 5 proveedores por facturación
    top_proveedores_result = await db.execute(
        select(
            models.Proveedor.id,
            models.Proveedor.nombre,
            func.sum(models.FacturaOficina.valor).label('total')
        )
        .join(models.Factura, models.Factura.proveedor_id == models.Proveedor.id)
        .join(models.FacturaOficina, models.FacturaOficina.factura_id == models.Factura.id)
        .filter(
            or_(
                and_(
                    models.Factura.fecha_factura.isnot(None),
                    models.Factura.fecha_factura >= start_date,
                    models.Factura.fecha_factura <= end_date
                ),
                and_(
                    models.Factura.fecha_factura.is_(None),
                    cast(models.Factura.created_at, Date) >= start_date,
                    cast(models.Factura.created_at, Date) <= end_date
                )
            )
        )
        .group_by(models.Proveedor.id, models.Proveedor.nombre)
        .order_by(func.sum(models.FacturaOficina.valor).desc())
        .limit(5)
    )
    top_proveedores = [
        {'id': row[0], 'nombre': row[1], 'total': float(row[2] or 0)}
        for row in top_proveedores_result.all()
    ]
    
    # 7. Facturación por tipo de contrato
    facturacion_por_tipo_result = await db.execute(
        select(
            models.Contrato.tipo,
            func.sum(models.FacturaOficina.valor).label('total')
        )
        .join(models.FacturaOficina, models.FacturaOficina.contrato_id == models.Contrato.id)
        .join(models.Factura, models.FacturaOficina.factura_id == models.Factura.id)
        .filter(
            or_(
                and_(
                    models.Factura.fecha_factura.isnot(None),
                    models.Factura.fecha_factura >= start_date,
                    models.Factura.fecha_factura <= end_date
                ),
                and_(
                    models.Factura.fecha_factura.is_(None),
                    cast(models.Factura.created_at, Date) >= start_date,
                    cast(models.Factura.created_at, Date) <= end_date
                )
            )
        )
        .group_by(models.Contrato.tipo)
    )
    facturacion_por_tipo = [
        {'tipo': row[0] or 'Sin tipo', 'total': float(row[1] or 0)}
        for row in facturacion_por_tipo_result.all()
    ]
    
    return {
        "año": target_year,
        "resumen": {
            "total_facturado": total_facturado,
            "total_facturas": total_facturas,
            "proveedores_facturados": proveedores_facturados,
            "contratos_activos": contratos_activos
        },
        "facturacion_mensual": facturacion_por_mes,
        "top_proveedores": top_proveedores,
        "facturacion_por_tipo": facturacion_por_tipo
    }


@router.get("/reportes/contratos-por-nit/{nit}")
async def get_contratos_by_nit(
    nit: str,
    db: AsyncSession = Depends(get_db)
):
    """
    Get all contracts for a provider by their NIT.
    Returns: num_contrato, valor, cod_oficina, nombre_oficina, direccion, ciudad
    """
    # First find the provider by NIT
    proveedor_result = await db.execute(
        select(models.Proveedor).filter(models.Proveedor.nit == nit)
    )
    proveedor = proveedor_result.scalar_one_or_none()
    
    if not proveedor:
        return {
            "error": "Proveedor no encontrado",
            "nit": nit,
            "contratos": []
        }
    
    # Get all contracts for this provider with office information
    contratos_result = await db.execute(
        select(models.Contrato)
        .options(selectinload(models.Contrato.oficina))
        .filter(models.Contrato.proveedor_id == proveedor.id)
        .order_by(models.Contrato.id)
    )
    contratos = contratos_result.scalars().all()
    
    # Build simplified response
    contratos_data = []
    for c in contratos:
        contratos_data.append({
            "num_contrato": c.num_contrato or "",
            "valor": float(c.valor_mensual) if c.valor_mensual else 0,
            "cod_oficina": c.oficina.cod_oficina if c.oficina else "",
            "nombre_oficina": c.oficina.nombre if c.oficina else "",
            "direccion": c.oficina.direccion if c.oficina else "",
            "ciudad": c.oficina.ciudad if c.oficina else ""
        })
    
    return {
        "nit": nit,
        "nombre_proveedor": proveedor.nombre,
        "total_contratos": len(contratos_data),
        "contratos": contratos_data
    }
