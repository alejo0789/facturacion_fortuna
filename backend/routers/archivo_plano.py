"""
Archivo Plano Router - Generate flat file Excel for Manager accounting system
"""
from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import Optional, List
from decimal import Decimal
from datetime import date, datetime
import io
import os
import httpx
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
from dotenv import load_dotenv

load_dotenv()

router = APIRouter()

# Path to template file
TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), '..', 'Template_archivo_plano', 'template_plano.xlsx')

# --- Schemas ---

class OficinaArchivoPlano(BaseModel):
    """Office with value for flat file generation"""
    cod_oficina: str
    valor: Decimal
    nombre_oficina: Optional[str] = None
    num_contrato: Optional[str] = None


class FacturaArchivoPlano(BaseModel):
    """Invoice with its offices for flat file generation"""
    id: Optional[int] = None
    numero_factura: Optional[str] = None
    fecha_factura: Optional[date] = None  # For extracting month
    oficinas: List[OficinaArchivoPlano]


class ArchivoPlanoRequest(BaseModel):
    """Request to generate flat file Excel"""
    proveedor_nit: str
    proveedor_nombre: Optional[str] = None
    fecha_causacion: Optional[date] = None  # Defaults to today
    tiene_iva: bool = True  # If supplier has IVA
    porcentaje_retefuente: float = 0  # 0 = no retefuente, 4 = 4%, 6 = 6%
    facturas: List[FacturaArchivoPlano]  # List of invoices with their offices
    numedoc: int = 1290  # Variable for future updates


# --- Helper Functions ---

def clean_oficina_code(cod_oficina: str) -> str:
    """
    Remove internal suffix from office code if present.
    Example: '001_INT_1' -> '001'
    """
    if not cod_oficina:
        return cod_oficina
    upper_cod = cod_oficina.upper()
    # REGLA DE EMERGENCIA: Si el código es de la serie 001, forzar 001
    if "001_INT" in upper_cod:
        return "001"
        
    if "_INT_" in upper_cod:
        # Split using the uppercase version to find the position, but slice the original
        idx = upper_cod.find("_INT_")
        res = cod_oficina[:idx]
        # REGLA ESPECIAL: Si lo que queda es '001', asegurar que no haya colas
        if res.strip() == "001":
            return "001"
        return res
    return cod_oficina


def extract_codigo_for_oracle(cod_oficina: str) -> str:
    """
    Extract digits to search Oracle based on cod_oficina length:
    - 7 digits -> first 4
    - 6 digits -> first 3
    - 5 digits -> first 2
    - 4 digits -> first 1
    """
    # Clean internal code first
    cod = clean_oficina_code(cod_oficina.strip())
    length = len(cod)
    
    if length >= 7:
        return cod[:4]
    elif length == 6:
        return cod[:3]
    elif length == 5:
        return cod[:2]
    elif length == 4:
        return cod[:1]
    else:
        return cod

    return ""


async def get_centro_costo(cod_oficina: str) -> str:
    """
    Call Oracle API to get centro de costo for an office.
    Returns the codigo_ccosto or empty string if not found.
    Tries with truncated code first, then falls back to full cleaned code.
    """
    api_key = os.getenv("API_KEY", "")
    
    # helper uses clean_oficina_code internally
    codigo_busqueda = extract_codigo_for_oracle(cod_oficina)
    codigo_completo = clean_oficina_code(cod_oficina.strip())
    
    # HARDCODED RULES for specific offices per user request
    if codigo_busqueda == "001":
        return "0401"
    if codigo_busqueda == "010":
        return "02"
    
    try:
        async with httpx.AsyncClient() as client:
            # First attempt: partial code
            response = await client.get(
                f"http://localhost:8000/api/oficinas-oracle/{codigo_busqueda}",
                headers={"X-API-Key": api_key},
                timeout=10.0
            )
            
            if response.status_code == 200:
                data = response.json()
                if data.get("success") and data.get("data"):
                    return data["data"].get("codigo_ccosto", "").strip()
            
            # Second attempt: full code (if different)
            if codigo_busqueda != codigo_completo:
                print(f"Retrying with full code for {cod_oficina}: {codigo_completo}")
                response = await client.get(
                    f"http://localhost:8000/api/oficinas-oracle/{codigo_completo}",
                    headers={"X-API-Key": api_key},
                    timeout=10.0
                )
                
                if response.status_code == 200:
                    data = response.json()
                    if data.get("success") and data.get("data"):
                        return data["data"].get("codigo_ccosto", "").strip()

    except Exception as e:
        print(f"Error getting centro costo for {cod_oficina}: {e}")
    
    return ""


def format_date_for_excel(d: date) -> str:
    """Format date as YYYY/MM/DD for Excel (template has text format)"""
    return d.strftime('%Y/%m/%d')


def format_value(value: str) -> str:
    """Return value as-is (template already has text format configured)"""
    return value


def get_month_name_spanish(d: date) -> str:
    """Get Spanish month name from date"""
    months = {
        1: "ENERO", 2: "FEBRERO", 3: "MARZO", 4: "ABRIL",
        5: "MAYO", 6: "JUNIO", 7: "JULIO", 8: "AGOSTO",
        9: "SEPTIEMBRE", 10: "OCTUBRE", 11: "NOVIEMBRE", 12: "DICIEMBRE"
    }
    return months.get(d.month, "")


def build_detalle(numero_factura: str, nombre_oficina: str, mes_factura: str, proveedor_nit: str, num_contrato: Optional[str] = None) -> str:
    """Helper to build consistent detail strings with special rules for certain NITs"""
    # Base observation (standard format)
    base = f"FACT {numero_factura} SERVICIO DE INTERNET {nombre_oficina} MES {mes_factura}"

    # NITs that include the contract number in the middle of the detail
    nit_especiales = ["830114921", "830122566", "800153993", "891502163", "91502163", "900092385"]
    if proveedor_nit in nit_especiales and num_contrato:
        base = f"FACT {numero_factura}, Contrato {num_contrato}, SERVICIO DE INTERNET {nombre_oficina} MES {mes_factura}"

    # NIT 900971687 (Hughes): prefix with "REF <num_contrato>,"
    if proveedor_nit == "900971687" and num_contrato:
        return f"REF {num_contrato}, {base}"

    return base


# --- Casos especiales: cuenta única (sin división 70/30) ---
# Estructura de cada regla:
#   "contratos": set de números de contrato específicos, o None para aplicar a TODOS los contratos
#   "cuenta": cuenta contable donde va el 100% del valor base
CUENTA_UNICA_REGLAS = {
    # NIT 830122566: solo en contratos 10434167091 y 181161832 → cuenta 51209505
    "830122566": {"contratos": {"10434167091", "181161832"}, "cuenta": "51209505"},
    # NIT 819006966 (Medicommerce): todos los contratos → cuenta 51353503
    "819006966": {"contratos": None, "cuenta": "51353503"},
}

def get_cuenta_unica(proveedor_nit: str, num_contrato: Optional[str]) -> Optional[str]:
    """
    Returns the single account to use (100% of valor_base) if this NIT+contrato
    combination requires it, or None if the normal 70/30 split applies.

    Rules:
    - If regla["contratos"] is None  -> applies to ALL contracts of that NIT.
    - If regla["contratos"] is a set -> applies only when num_contrato is in that set.
    """
    regla = CUENTA_UNICA_REGLAS.get(proveedor_nit)
    if not regla:
        return None
    contratos_especificos = regla["contratos"]
    if contratos_especificos is None:
        # Applies to all contracts (and even when there is no contract)
        return regla["cuenta"]
    if num_contrato and num_contrato.strip() in contratos_especificos:
        return regla["cuenta"]
    return None


def create_flat_file_row(
    row_index: int,  # Excel row number (2, 3, 4...) for formulas
    empresa: str = "101",  # Must be text, not number
    clase: str = "0000 ",
    vinkey: str = ".",
    tipodoc: str = "DC07",  # Template handles text format
    numedoc: int = 1290,
    reg: int = 0,  # Changed to 0
    fecha: str = "",
    cuenta: str = "",
    vinculado: str = "",
    sucvin: str = ".",
    sucurs: str = ".",
    ccosto: str = "",
    destino: str = "",
    vende: str = ".",
    cobra: str = ".",
    zona: str = ".",
    bodega: str = ".",
    producto: str = ".",
    unimed: str = ".",
    lotepro: str = ".",
    cantidad: int = 0,  # Changed to 0
    claseinv: str = ".",
    clacru1: str = "0000 ",
    tipcru1: str = None,  # Will be formula =D{row}
    numcru1: str = None,   # Will be formula =E{row}
    cuocru1: int = 0,     # Changed to 0
    fecini: str = None,    # Will be formula =G{row}
    plazo: int = 0,       # Changed to 0
    clacru2: str = ".",
    tipcru2: str = ".",
    numcru2: int = 0,     # Changed to 0
    cuocru2: int = 0,     # Changed to 0
    valdebi: float = 0,   # Always 0, not empty
    valcred: float = 0,   # Always 0, not empty
    parci_o: int = 0,     # Changed to 0
    tpreg: int = 1,       # Must be number, not text
    detalle: str = "",
    serial: str = ".",
    formapago: str = ".",
    dv_referencia: str = ".",
    dv_motivo: str = ".",
    docrespald: str = None,  # Will be formula =E{row}
    docplazo: int = 0        # Changed to 0
) -> list:
    """Create a single row for the flat file with Excel formulas"""
    return [
        empresa,                       # EMPRESA.C3
        clase,                         # CLASE.C4
        vinkey,                        # VINKEY.C15
        tipodoc,                       # TIPODOC.C4 (with apostrophe)
        numedoc,                       # NUMEDOC.N12
        reg,                           # REG.N12 = 0
        fecha,                         # FECHA.C10
        cuenta,                        # CUENTA.C14
        vinculado,                     # VINCULADO.C15
        sucvin,                        # SUCVIN.C3
        sucurs,                        # SUCURS.C5
        ccosto,                        # CCOSTO.C10
        destino,                       # DESTINO.C10
        vende,                         # VENDE.C5
        cobra,                         # COBRA.C5
        zona,                          # ZONA.C5
        bodega,                        # BODEGA.C5
        producto,                      # PRODUCTO.C15
        unimed,                        # UNIMED.C4
        lotepro,                       # LOTEPRO.C12
        cantidad,                      # CANTIDAD.N20 = 0
        claseinv,                      # CLASEINV.C1
        clacru1,                       # CLACRU1.C4
        tipcru1 if tipcru1 else f"=D{row_index}",  # TIPCRU1.C4 = formula
        numcru1 if numcru1 else f"=E{row_index}",  # NUMCRU1.N12 = formula
        cuocru1,                       # CUOCRU1.N12 = 0
        fecini if fecini else f"=G{row_index}",   # FECINI.C10 = formula
        plazo,                         # PLAZO.N10 = 0
        clacru2,                       # CLACRU2.C4
        tipcru2,                       # TIPCRU2.C4
        numcru2,                       # NUMCRU2.N12 = 0
        cuocru2,                       # CUOCRU2.N12 = 0
        valdebi,                       # VALDEBI.N20 (always number, 0 if no value)
        valcred,                       # VALCRED.N20 (always number, 0 if no value)
        parci_o,                       # PARCI_O.N20 = 0
        tpreg,                         # TPREG.N1
        detalle,                       # DETALLE.C250
        serial,                        # SERIAL.C50
        formapago,                     # FORMAPAGO.C10
        dv_referencia,                 # DV_REFERENCIA.C80
        dv_motivo,                     # DV_MOTIVO.C6
        docrespald if docrespald else f"=E{row_index}",  # DOCRESPALD.N15 = formula
        docplazo                       # DOCPLAZO.N15 = 0
    ]


async def generate_rows_for_oficina(
    oficina: OficinaArchivoPlano,
    proveedor_nit: str,
    fecha: str,
    numedoc: int,
    tiene_iva: bool,
    numero_factura: str,  # For building DETALLE
    fecha_factura: Optional[date],  # For extracting month
    starting_row_index: int  # Excel row number to start (2, 3, 4...)
) -> tuple[List[list], int, dict]:
    """
    Generate debit rows for a single office.

    Normal case: splits valor_base 70% (cuenta 61350513) + 30% (cuenta 61700360).
    Special case: NIT+contrato in CUENTA_UNICA_REGLAS → 100% to a single account (e.g. 51209505).

    The base is calculated on (valor / 1.19) if tiene_iva.

    Returns:
        tuple: (list of rows, next_row_index, office_info for final rows)
    """
    rows = []
    current_row = starting_row_index

    # Get centro de costo from Oracle
    ccosto_raw = await get_centro_costo(oficina.cod_oficina)
    ccosto = format_value(ccosto_raw) if ccosto_raw else ""

    # Format values
    vinculado = format_value(proveedor_nit)
    destino = format_value(clean_oficina_code(oficina.cod_oficina))

    # Build DETALLE
    nombre_oficina = oficina.nombre_oficina or oficina.cod_oficina
    mes_factura = get_month_name_spanish(fecha_factura) if fecha_factura else ""

    detalle = build_detalle(
        numero_factura=numero_factura,
        nombre_oficina=nombre_oficina,
        mes_factura=mes_factura,
        proveedor_nit=proveedor_nit,
        num_contrato=oficina.num_contrato
    )

    valor = round(float(oficina.valor), 0)  # Valor total de la oficina (ENTERO)

    # Calculate base value (without IVA if applicable)
    if tiene_iva:
        valor_base = round(valor / 1.19, 0)
        valor_iva = round(valor - valor_base, 0)
    else:
        valor_base = valor
        valor_iva = 0

    # Check if this NIT+contrato uses a single account instead of 70/30 split
    cuenta_unica = get_cuenta_unica(proveedor_nit, oficina.num_contrato)

    if cuenta_unica:
        # Special case: 100% of valor_base to a single account
        rows.append(create_flat_file_row(
            row_index=current_row,
            numedoc=numedoc,
            fecha=fecha,
            cuenta=format_value(cuenta_unica),
            vinculado=vinculado,
            ccosto=ccosto,
            destino=destino,
            valdebi=valor_base,
            detalle=detalle
        ))
        current_row += 1
        valor_70 = valor_base  # For accumulator compatibility
        valor_30 = 0
    else:
        # Normal case: 70% + 30% split
        valor_70 = round(valor_base * 0.70, 0)
        valor_30 = round(valor_base - valor_70, 0)

        # Row 1: Account 61350513 - 70% (VALDEBI)
        rows.append(create_flat_file_row(
            row_index=current_row,
            numedoc=numedoc,
            fecha=fecha,
            cuenta=format_value("61350513"),
            vinculado=vinculado,
            ccosto=ccosto,
            destino=destino,
            valdebi=valor_70,
            detalle=detalle
        ))
        current_row += 1

        # Row 2: Account 61700360 - 30% (VALDEBI)
        rows.append(create_flat_file_row(
            row_index=current_row,
            numedoc=numedoc,
            fecha=fecha,
            cuenta=format_value("61700360"),
            vinculado=vinculado,
            ccosto=ccosto,
            destino=destino,
            valdebi=valor_30,
            detalle=detalle
        ))
        current_row += 1

    # Return office info for final rows
    office_info = {
        "ccosto": ccosto,
        "destino": destino,
        "vinculado": vinculado,
        "valor": valor,
        "valor_base": valor_base,
        "valor_iva": valor_iva,
        "valor_70": valor_70,
        "valor_30": valor_30,
        "detalle": detalle,
        "num_contrato": oficina.num_contrato
    }

    return rows, current_row, office_info


def create_final_summary_rows(
    total_debitos: float,  # Sum of all 70% + 30% across all offices
    total_iva: float,      # Total IVA across all offices
    tiene_iva: bool,
    porcentaje_retefuente: float,  # 0, 4, or 6
    total_valor_base: float,    # Sum of valor_base (SIN IVA) for retefuente calculation
    last_office_info: dict,  # ccosto, destino, vinculado from last office
    numedoc: int,
    fecha: str,
    detalle: str,
    starting_row_index: int
) -> tuple[List[list], int]:
    """
    Generate final summary rows: IVA, Retefuente (if applicable), and Total.
    Uses the last office's ccosto and destino.
    """
    rows = []
    current_row = starting_row_index
    
    ccosto = last_office_info["ccosto"]
    destino = last_office_info["destino"]
    vinculado = last_office_info["vinculado"]
    
    # Calculate retefuente based on percentage (0%, 4%, or 6%) - SOBRE VALOR BASE SIN IVA
    # Round to integer as requested
    valor_retefuente = round(total_valor_base * (porcentaje_retefuente / 100), 0) if porcentaje_retefuente > 0 else 0
    
    # Calculate balance: total debitos + IVA - retefuente
    # Calculations are already integers or rounded to integers
    valor_balance = total_debitos + total_iva - valor_retefuente
    
    # Row: Account 24081003 - IVA total (VALDEBI) - only if tiene_iva
    if tiene_iva and total_iva > 0:
        rows.append(create_flat_file_row(
            row_index=current_row,
            numedoc=numedoc,
            fecha=fecha,
            cuenta=format_value("24081003"),
            vinculado=vinculado,
            ccosto=ccosto,
            destino=".",  # IVA row uses "." for DESTINO
            valdebi=total_iva,
            detalle=detalle
        ))
        current_row += 1
    
    # Row: Account 23652501 - Retefuente (VALCRED) - only if porcentaje > 0
    if porcentaje_retefuente > 0:
        rows.append(create_flat_file_row(
            row_index=current_row,
            numedoc=numedoc,
            fecha=fecha,
            cuenta=format_value("23652501"),
            vinculado=vinculado,
            ccosto=ccosto,
            destino=destino,
            valcred=valor_retefuente,
            detalle=detalle
        ))
        current_row += 1
    
    # Row: Account 23355002 - Balance total (VALCRED)
    rows.append(create_flat_file_row(
        row_index=current_row,
        numedoc=numedoc,
        fecha=fecha,
        cuenta=format_value("23355002"),
        vinculado=vinculado,
        ccosto=ccosto,
        destino=destino,
        valcred=valor_balance,
        detalle=detalle
    ))
    current_row += 1
    
    return rows, current_row


# --- Constants: Column Headers ---
HEADERS = [
    "EMPRESA.C3", "CLASE.C4", "VINKEY.C15", "TIPODOC.C4", "NUMEDOC.N12",
    "REG.N12", "FECHA.C10", "CUENTA.C14", "VINCULADO.C15", "SUCVIN.C3",
    "SUCURS.C5", "CCOSTO.C10", "DESTINO.C10", "VENDE.C5", "COBRA.C5",
    "ZONA.C5", "BODEGA.C5", "PRODUCTO.C15", "UNIMED.C4", "LOTEPRO.C12",
    "CANTIDAD.N20", "CLASEINV.C1", "CLACRU1.C4", "TIPCRU1.C4", "NUMCRU1.N12",
    "CUOCRU1.N12", "FECINI.C10", "PLAZO.N10", "CLACRU2.C4", "TIPCRU2.C4",
    "NUMCRU2.N12", "CUOCRU2.N12", "VALDEBI.N20", "VALCRED.N20", "PARCI_O.N20",
    "TPREG.N1", "DETALLE.C250", "SERIAL.C50", "FORMAPAGO.C10",
    "DV_REFERENCIA.C80", "DV_MOTIVO.C6", "DOCRESPALD.N15", "DOCPLAZO.N15"
]


# --- Endpoint ---

@router.post("/archivo-plano/generar")
async def generar_archivo_plano(request: ArchivoPlanoRequest):
    """
    Generate flat file Excel for Manager accounting system.
    
    For each office, generates rows for:
    - Account 61350513: 70% of value (debit)
    - Account 61700360: 30% of value (debit)
    - Account 24081003: IVA 19% (debit) - if tiene_iva
    - Account 23652501: Retefuente (credit) - if porcentaje_retefuente > 0
    - Account 23355002: Balance total (credit)
    """
    if not request.facturas:
        raise HTTPException(status_code=400, detail="Debe proporcionar al menos una factura")
    
    # Use today's date if not provided
    fecha_causacion = request.fecha_causacion or date.today()
    fecha_str = format_date_for_excel(fecha_causacion)
    
    # Generate all rows
    all_rows = []
    current_row_index = 2  # Excel rows start at 2 (row 1 is headers)
    
    # Process each factura separately with incrementing NUMEDOC
    for factura_index, factura in enumerate(request.facturas):
        if not factura.oficinas:
            continue
        
        # NUMEDOC increments per factura: first factura uses base, next uses base+1, etc.
        factura_numedoc = request.numedoc + factura_index
        
        # Accumulators for this factura
        factura_debitos = 0  # Sum of 70% + 30%
        factura_iva = 0
        factura_valor_base = 0  # Sum of valor_base (sin IVA) para calcular retefuente
        last_office_info = None
        last_detalle = ""  # Store the last detalle for summary rows
        
        # Generate rows for each office in this factura
        for oficina in factura.oficinas:
            rows, current_row_index, office_info = await generate_rows_for_oficina(
                oficina=oficina,
                proveedor_nit=request.proveedor_nit,
                fecha=fecha_str,
                numedoc=factura_numedoc,
                tiene_iva=request.tiene_iva,
                numero_factura=factura.numero_factura or '',
                fecha_factura=factura.fecha_factura,
                starting_row_index=current_row_index
            )
            all_rows.extend(rows)
            
            # Accumulate totals for this factura
            factura_debitos += office_info["valor_70"] + office_info["valor_30"]
            factura_iva += office_info["valor_iva"]
            factura_valor_base += office_info["valor_base"]  # Usar valor_base para retefuente
            last_office_info = office_info
            last_detalle = office_info.get("detalle", "")
        
        # Generate IVA and Total rows for THIS factura
        if last_office_info:
            summary_rows, current_row_index = create_final_summary_rows(
                total_debitos=factura_debitos,
                total_iva=factura_iva,
                tiene_iva=request.tiene_iva,
                porcentaje_retefuente=request.porcentaje_retefuente,
                total_valor_base=factura_valor_base,  # Pasar valor_base para retefuente
                last_office_info=last_office_info,
                numedoc=factura_numedoc,
                fecha=fecha_str,
                detalle=last_detalle,
                starting_row_index=current_row_index
            )
            all_rows.extend(summary_rows)
    
    # Load Excel template (preserves all cell formats)
    try:
        wb = load_workbook(TEMPLATE_PATH)
        ws = wb.active
    except FileNotFoundError:
        raise HTTPException(status_code=500, detail="Template file not found")
    
    # Write data rows (row 1 is headers in template, data starts at row 2)
    for row_idx, row_data in enumerate(all_rows, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
    
    # Clear excess rows (rows after our data that have old template data)
    last_data_row = len(all_rows) + 1  # +1 because data starts at row 2
    for row_idx in range(last_data_row + 1, ws.max_row + 1):
        for col_idx in range(1, 44):  # 43 columns
            ws.cell(row=row_idx, column=col_idx).value = None
    
    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # Generate filename: PLANO-PROVEEDOR-MES.xlsx
    mes_nombre = get_month_name_spanish(fecha_causacion)
    proveedor_slug = (request.proveedor_nombre or request.proveedor_nit).upper().replace(" ", "-")
    filename = f"PLANO-{proveedor_slug}-{mes_nombre}.xlsx"
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.post("/archivo-plano/preview")
async def preview_archivo_plano(request: ArchivoPlanoRequest):
    """
    Preview the flat file data without generating Excel.
    Returns JSON with all rows that would be generated.
    """
    if not request.facturas:
        raise HTTPException(status_code=400, detail="Debe proporcionar al menos una factura")
    
    fecha_causacion = request.fecha_causacion or date.today()
    fecha_str = format_date_for_excel(fecha_causacion)
    
    all_rows = []
    current_row_index = 2  # Excel rows start at 2 (row 1 is headers)
    
    # Process each factura separately with incrementing NUMEDOC
    for factura_index, factura in enumerate(request.facturas):
        if not factura.oficinas:
            continue
        
        # NUMEDOC increments per factura
        factura_numedoc = request.numedoc + factura_index
        
        # Accumulators for this factura
        factura_debitos = 0
        factura_iva = 0
        factura_valor = 0
        last_office_info = None
        last_detalle = ""
        
        for oficina in factura.oficinas:
            rows, current_row_index, office_info = await generate_rows_for_oficina(
                oficina=oficina,
                proveedor_nit=request.proveedor_nit,
                fecha=fecha_str,
                numedoc=factura_numedoc,
                tiene_iva=request.tiene_iva,
                numero_factura=factura.numero_factura or '',
                fecha_factura=factura.fecha_factura,
                starting_row_index=current_row_index
            )
            all_rows.extend(rows)
            
            factura_debitos += office_info["valor_70"] + office_info["valor_30"]
            factura_iva += office_info["valor_iva"]
            factura_valor += office_info["valor"]
            last_office_info = office_info
            last_detalle = office_info.get("detalle", "")
        
        # Generate IVA and Total rows for THIS factura
        if last_office_info:
            summary_rows, current_row_index = create_final_summary_rows(
                total_debitos=factura_debitos,
                total_iva=factura_iva,
                tiene_iva=request.tiene_iva,
                porcentaje_retefuente=request.porcentaje_retefuente,
                total_valor=factura_valor,
                last_office_info=last_office_info,
                numedoc=factura_numedoc,
                fecha=fecha_str,
                detalle=last_detalle,
                starting_row_index=current_row_index
            )
            all_rows.extend(summary_rows)
    
    # Convert rows to dict for better readability
    rows_as_dicts = []
    for row in all_rows:
        row_dict = {HEADERS[i]: row[i] for i in range(len(HEADERS))}
        rows_as_dicts.append(row_dict)
    
    return {
        "success": True,
        "total_rows": len(all_rows),
        "headers": HEADERS,
        "rows": rows_as_dicts
    }


# --- Manager Causation Preview Endpoint ---

class CausacionManagerPreviewRequest(BaseModel):
    """Request to preview causation data for Manager"""
    proveedor_nit: str
    proveedor_nombre: Optional[str] = None
    fecha_causacion: Optional[date] = None
    tiene_iva: bool = True
    porcentaje_retefuente: float = 0
    facturas: List[FacturaArchivoPlano]
    numedoc: int = 1290


class CausacionRowPreview(BaseModel):
    """A single row preview for the causation table"""
    row_num: int
    cuenta: str
    tipo_movimiento: str  # DEBITO or CREDITO
    ccosto: str
    destino: str
    valor: float
    detalle: str


class CausacionFacturaPreview(BaseModel):
    """Preview for a single factura"""
    id: Optional[int] = None
    numero_factura: str
    numedoc: int
    rows: List[CausacionRowPreview]
    total_debitos: float
    total_creditos: float


class CausacionManagerPreviewResponse(BaseModel):
    """Response with preview data for Manager causation"""
    success: bool
    proveedor_nit: str
    proveedor_nombre: Optional[str]
    fecha_causacion: str
    tiene_iva: bool
    porcentaje_retefuente: float
    facturas: List[CausacionFacturaPreview]
    total_facturas: int
    total_debitos: float
    total_creditos: float
    balance: float
    numedoc_inicial: int
    numedoc_final: int


@router.post("/causacion-manager/preview", response_model=CausacionManagerPreviewResponse)
async def preview_causacion_manager(request: CausacionManagerPreviewRequest):
    """
    Preview the causation data that would be sent to Manager.
    Returns a structured response for displaying in a table format.
    
    This does NOT insert anything into Manager - it's just a preview.
    """
    if not request.facturas:
        raise HTTPException(status_code=400, detail="Debe proporcionar al menos una factura")
    
    fecha_causacion = request.fecha_causacion or date.today()
    fecha_str = format_date_for_excel(fecha_causacion)
    
    facturas_preview: List[CausacionFacturaPreview] = []
    total_debitos_global = 0
    total_creditos_global = 0
    
    for factura_index, factura in enumerate(request.facturas):
        if not factura.oficinas:
            continue
        
        factura_numedoc = request.numedoc + factura_index
        rows_preview: List[CausacionRowPreview] = []
        factura_debitos = 0
        factura_creditos = 0
        factura_iva = 0
        factura_valor_base = 0  # Sum of valor_base (sin IVA) para calcular retefuente
        row_counter = 1
        
        # Process each oficina
        for oficina in factura.oficinas:
            ccosto_raw = await get_centro_costo(oficina.cod_oficina)
            ccosto = ccosto_raw if ccosto_raw else ""
            destino = clean_oficina_code(oficina.cod_oficina)
            nombre_oficina = oficina.nombre_oficina or oficina.cod_oficina
            mes_factura = get_month_name_spanish(factura.fecha_factura) if factura.fecha_factura else ""
            
            detalle = build_detalle(
                numero_factura=factura.numero_factura or '',
                nombre_oficina=nombre_oficina,
                mes_factura=mes_factura,
                proveedor_nit=request.proveedor_nit,
                num_contrato=oficina.num_contrato
            )
            
            valor = round(float(oficina.valor), 0)

            # Calculate base value
            if request.tiene_iva:
                valor_base = round(valor / 1.19, 0)
                valor_iva = round(valor - valor_base, 0)
            else:
                valor_base = valor
                valor_iva = 0

            # Check special single-account rule
            cuenta_unica = get_cuenta_unica(request.proveedor_nit, oficina.num_contrato)

            if cuenta_unica:
                # Special case: 100% to single account
                rows_preview.append(CausacionRowPreview(
                    row_num=row_counter,
                    cuenta=cuenta_unica,
                    tipo_movimiento="DEBITO",
                    ccosto=ccosto,
                    destino=destino,
                    valor=valor_base,
                    detalle=detalle
                ))
                row_counter += 1
                factura_debitos += valor_base
            else:
                # Normal 70/30 split
                valor_70 = round(valor_base * 0.70, 0)
                valor_30 = round(valor_base - valor_70, 0)

                # Row 1: Account 61350513 - 70% DEBITO
                rows_preview.append(CausacionRowPreview(
                    row_num=row_counter,
                    cuenta="61350513",
                    tipo_movimiento="DEBITO",
                    ccosto=ccosto,
                    destino=destino,
                    valor=valor_70,
                    detalle=detalle
                ))
                row_counter += 1
                factura_debitos += valor_70

                # Row 2: Account 61700360 - 30% DEBITO
                rows_preview.append(CausacionRowPreview(
                    row_num=row_counter,
                    cuenta="61700360",
                    tipo_movimiento="DEBITO",
                    ccosto=ccosto,
                    destino=destino,
                    valor=valor_30,
                    detalle=detalle
                ))
                row_counter += 1
                factura_debitos += valor_30
            
            factura_iva += valor_iva
            factura_valor_base += valor_base  # Acumular valor_base para retefuente
        
        # Get last office info for summary rows
        if factura.oficinas:
            last_oficina = factura.oficinas[-1]
            last_ccosto_raw = await get_centro_costo(last_oficina.cod_oficina)
            last_ccosto = last_ccosto_raw if last_ccosto_raw else ""
            last_destino = clean_oficina_code(last_oficina.cod_oficina)
            last_nombre = last_oficina.nombre_oficina or last_oficina.cod_oficina
            last_mes = get_month_name_spanish(factura.fecha_factura) if factura.fecha_factura else ""
            last_detalle = build_detalle(
                numero_factura=factura.numero_factura or '',
                nombre_oficina=last_nombre,
                mes_factura=last_mes,
                proveedor_nit=request.proveedor_nit,
                num_contrato=last_oficina.num_contrato
            )
        
        # IVA row (DEBITO)
        if request.tiene_iva and factura_iva > 0:
            rows_preview.append(CausacionRowPreview(
                row_num=row_counter,
                cuenta="24081003",
                tipo_movimiento="DEBITO",
                ccosto=last_ccosto,
                destino=".",
                valor=factura_iva,
                detalle=last_detalle
            ))
            row_counter += 1
            factura_debitos += factura_iva
        
        # Retefuente row (CREDITO) - SOBRE VALOR BASE SIN IVA
        valor_retefuente = round(factura_valor_base * (request.porcentaje_retefuente / 100), 0) if request.porcentaje_retefuente > 0 else 0
        if valor_retefuente > 0:
            rows_preview.append(CausacionRowPreview(
                row_num=row_counter,
                cuenta="23652501",
                tipo_movimiento="CREDITO",
                ccosto=last_ccosto,
                destino=last_destino,
                valor=valor_retefuente,
                detalle=last_detalle
            ))
            row_counter += 1
            factura_creditos += valor_retefuente
        
        # Balance row (CREDITO) - Total to pay supplier
        valor_balance = factura_debitos - factura_creditos
        rows_preview.append(CausacionRowPreview(
            row_num=row_counter,
            cuenta="23355002",
            tipo_movimiento="CREDITO",
            ccosto=last_ccosto,
            destino=last_destino,
            valor=valor_balance,
            detalle=last_detalle
        ))
        factura_creditos += valor_balance
        
        facturas_preview.append(CausacionFacturaPreview(
            id=factura.id,
            numero_factura=factura.numero_factura or f"Factura {factura_index + 1}",
            numedoc=factura_numedoc,
            rows=rows_preview,
            total_debitos=factura_debitos,
            total_creditos=factura_creditos
        ))
        
        total_debitos_global += factura_debitos
        total_creditos_global += factura_creditos
    
    numedoc_final = request.numedoc + len([f for f in request.facturas if f.oficinas]) - 1
    
    return CausacionManagerPreviewResponse(
        success=True,
        proveedor_nit=request.proveedor_nit,
        proveedor_nombre=request.proveedor_nombre,
        fecha_causacion=fecha_str,
        tiene_iva=request.tiene_iva,
        porcentaje_retefuente=request.porcentaje_retefuente,
        facturas=facturas_preview,
        total_facturas=len(facturas_preview),
        total_debitos=total_debitos_global,
        total_creditos=total_creditos_global,
        balance=total_debitos_global - total_creditos_global,
        numedoc_inicial=request.numedoc,
        numedoc_final=numedoc_final
    )


# --- Manager Causation INSERT Endpoint ---

class CausacionInsertRequest(BaseModel):
    """Request to insert causation data into Manager"""
    proveedor_nit: str
    proveedor_nombre: Optional[str] = None
    fecha_causacion: Optional[date] = None
    tiene_iva: bool = True
    porcentaje_retefuente: float = 0
    facturas: List[CausacionFacturaPreview]
    numedoc: int


class CausacionInsertResponse(BaseModel):
    """Response from causation insert"""
    success: bool
    message: str
    numedoc_inicial: int
    numedoc_final: int
    total_registros_mngdoc: int
    total_registros_mngmcn: int
    error: Optional[str] = None


@router.post("/causacion-manager/insertar", response_model=CausacionInsertResponse)
async def insertar_causacion_manager(request: CausacionInsertRequest):
    """
    Insert causation data into Manager ERP.
    
    This endpoint inserts:
    1. One record per factura into MNGDOC (header)
    2. Multiple records per factura into MNGMCN (details)
    
    The insert is done in a transaction - if any insert fails, all are rolled back.
    """
    import sys
    sys.path.append('..')
    from oracle_database import get_oracle_connection
    
    if not request.facturas:
        raise HTTPException(status_code=400, detail="Debe proporcionar al menos una factura")
    
    fecha_causacion = request.fecha_causacion or date.today()
    fecha_str = fecha_causacion.strftime('%Y-%m-%d')
    
    connection = None
    cursor = None
    total_mngdoc = 0
    total_mngmcn = 0
    
    try:
        connection = get_oracle_connection()
        cursor = connection.cursor()
        
        for factura_index, factura in enumerate(request.facturas):
            if not factura.rows:
                continue
            
            factura_numedoc = request.numedoc + factura_index
            
            # Get data for header from the first row of the preview
            first_row = factura.rows[0]
            ccosto = first_row.ccosto if first_row.ccosto else "."
            destino = first_row.destino if first_row.destino else "."
            detalle_cabecera = first_row.detalle if first_row.detalle else f"FACT {factura.numero_factura}"
            
            # === INSERT INTO MNGDOC (Header) ===
            cursor.execute("""
                INSERT INTO MANAGER.MNGDOC (
                    DOCEMPRESA, DOCCLASE, DOCVINKEY, DOCTIPO, DOCNUMERO,
                    DOCSUCURS, DOCFECHA, DOCVINCULA, DOCSUCVIN, DOCCCOSTO,
                    DOCDESTINO, DOCLOTE, DOCVENDE, DOCZONA, DOCCOBRA,
                    DOCRESPALD, DOCPOSTFEC, DOCNEWUSER, DOCNEWFEC, DOCMODUSER,
                    DOCMODFEC, DOCPLAZOD, DOCESTADO, DOCRESPAL2, DOCNIMPRE,
                    DOCCODEU_1, DOCCODEU_2, DOCFORPAGO, DOCTARIFA, DOCBOD1E,
                    DOCBOD2S, DOCINTERES, DOCFECHA2, DOCPRODUCT, DOCCANTI,
                    DOCUNIMED, DOCRESPAL3, DOCNOTA2, DOCDETALLE
                ) VALUES (
                    '101', '0000', '.', 'DC07', :numedoc,
                    '.', TO_DATE(:fecha, 'YYYY-MM-DD'), :nit, '.', :ccosto,
                    :destino, '.', '.', '.', '.',
                    :numedoc, TO_DATE(:fecha, 'YYYY-MM-DD'), 'WEBAPP', SYSDATE, 'WEBAPP',
                    SYSDATE, 0, 'a', 0, 0,
                    '.', '.', '.', 1, '.',
                    '.', 0, TO_DATE(:fecha, 'YYYY-MM-DD'), '.', 0,
                    '.', ' ', NULL, :detalle
                )
            """, {
                'numedoc': factura_numedoc,
                'fecha': fecha_str,
                'nit': request.provider_nit if hasattr(request, 'provider_nit') else request.proveedor_nit,
                'ccosto': ccosto,
                'destino': destino,
                'detalle': detalle_cabecera[:2000]
            })
            total_mngdoc += 1
            
            # === Process rows exactly as received from preview ===
            for row in factura.rows:
                debito = row.valor if row.tipo_movimiento == "DEBITO" else 0
                credito = row.valor if row.tipo_movimiento == "CREDITO" else 0
                
                # Check if it is the counterparty row (23355002) for saldo
                saldocr = credito if str(row.cuenta).strip() == "23355002" else 0

                cursor.execute("""
                    INSERT INTO MANAGER.MNGMCN (
                        MCNEMPRESA, MCNCLASE, MCNVINKEY, MCNTIPODOC, MCNNUMEDOC, MCNREG, MCNFECHA,
                        MCNCLACRU1, MCNTIPCRU1, MCNNUMCRU1, MCNCUOCRU1, MCNSUCURS, MCNCUENTA, MCNVINCULA,
                        MCNSUCVIN, MCNCCOSTO, MCNDESTINO, MCNVENDE, MCNCOBRA, MCNZONA, MCNFECINI, MCNPLAZO,
                        MCNVALDEBI, MCNVALCRED, MCNTASA, MCNBASE, MCNCLACRU2, MCNTIPCRU2, MCNNUMCRU2, MCNCUOCRU2,
                        MCNSALDODB, MCNSALDOCR, MCNNEWUSER, MCNNEWFEC, MCNMODUSER, MCNMODFEC, MCNBODEGA,
                        MCNPROPADR, MCNPRODUCT, MCNCANTI_O, MCNUNI_O, MCNPARCI_O, MCNCANTID, MCNUNIDAD,
                        MCNPRECIOB, MCNFACTOR, MCNDCTO1, MCNDCTO2, MCNDCTO3, MCNDCTO4, MCNIMPOCON, MCNPRCOSVT,
                        MCNIVATIPO, MCNIVAPORC, MNCNIVAINC, MCNCOSTORE, MCNDIMEORI, MCNINDINV, MCNLOTEPRO,
                        MCNPRECIOX, MCNREF1, MCNREF2, MCNESTADO, MCNDETALLE, MCNFTE, MCNTPREG
                    ) VALUES (
                        '101', '0000', '.', 'DC07', :numedoc, :reg, TO_DATE(:fecha, 'YYYY-MM-DD'),
                        '0000', 'DC07', :numedoc, 0, '.', :cuenta, :nit,
                        '.', :ccosto, :destino, '.', '.', '.', TO_DATE(:fecha, 'YYYY-MM-DD'), 0,
                        :valdebi, :valcred, 0, 0, ' ', ' ', 0, 0,
                        0, :saldocr, 'WEBAPP', SYSDATE, 'WEBAPP', SYSDATE, '.',
                        '.', '.', 0, '.', 0, 0, '.',
                        0, 1, 0, 0, 0, 0, 0, 0,
                        '.', 0, 0, 0, 0, '.', '.',
                        0, '.', '.', 'a', :detalle, '.', 1
                    )
                """, {
                    'numedoc': factura_numedoc,
                    'reg': row.row_num,
                    'fecha': fecha_str,
                    'nit': request.proveedor_nit,
                    'cuenta': row.cuenta,
                    'ccosto': row.ccosto if row.ccosto else ".",
                    'destino': row.destino if row.destino else ".",
                    'valdebi': debito,
                    'valcred': credito,
                    'saldocr': saldocr,
                    'detalle': row.detalle[:4000]
                })
                total_mngmcn += 1
        
        # Commit all changes to Oracle
        connection.commit()
        
        # Now update our local database to save the Documento Contable (DC07-XXX)
        try:
            from database import SessionLocal
            from models import Factura
            import re
            from sqlalchemy import select
            
            async with SessionLocal() as db_local:
                for factura_index, factura in enumerate(request.facturas):
                    if factura.id:
                        f_numedoc = request.numedoc + factura_index
                        doc_str = f"DC07-{f_numedoc}"
                        
                        stmt = select(Factura).where(Factura.id == factura.id)
                        res = await db_local.execute(stmt)
                        f_db = res.scalar_one_or_none()
                        
                        if f_db:
                            # Update observations with the ACTUAL detail from the preview
                            f_db.observaciones = f"{detalle_cabecera} | Ref Doc: {doc_str}"
                                    
                            # Update status to EN TRAMITE
                            f_db.estado = "EN TRAMITE"
                            
                await db_local.commit()
        except Exception as local_db_error:
            print(f"Warning: Failed to update local DB with documento_contable/estado: {local_db_error}")
        
        numedoc_final = request.numedoc + len(request.facturas) - 1
        
        return CausacionInsertResponse(
            success=True,
            message=f"Causación insertada exitosamente. NUMEDOC: {request.numedoc} - {numedoc_final}",
            numedoc_inicial=request.numedoc,
            numedoc_final=numedoc_final,
            total_registros_mngdoc=total_mngdoc,
            total_registros_mngmcn=total_mngmcn
        )
        
    except Exception as e:
        # Rollback on error
        if connection:
            connection.rollback()
        return CausacionInsertResponse(
            success=False,
            message="Error al insertar causación",
            numedoc_inicial=request.numedoc,
            numedoc_final=request.numedoc,
            total_registros_mngdoc=0,
            total_registros_mngmcn=0,
            error=str(e)
        )
    finally:
        if cursor:
            cursor.close()
        if connection:
            connection.close()


# --- Diagnostic Endpoint: Inspect MNGMCN table structure ---

@router.get("/mngmcn/estructura")
async def get_mngmcn_estructura():
    """
    Diagnostic endpoint to get the structure of MANAGER.MNGMCN table.
    Returns column names, data types, and sample data.
    """
    import sys
    sys.path.append('..')
    from oracle_database import get_oracle_connection
    
    connection = None
    cursor = None
    try:
        connection = get_oracle_connection()
        cursor = connection.cursor()
        
        # Get table structure
        cursor.execute("""
            SELECT 
                COLUMN_NAME,
                DATA_TYPE,
                DATA_LENGTH,
                DATA_PRECISION,
                DATA_SCALE,
                NULLABLE
            FROM ALL_TAB_COLUMNS 
            WHERE OWNER = 'MANAGER' 
            AND TABLE_NAME = 'MNGMCN'
            ORDER BY COLUMN_ID
        """)
        columns = cursor.fetchall()
        
        estructura = []
        for col in columns:
            estructura.append({
                "column_name": col[0],
                "data_type": col[1],
                "data_length": col[2],
                "data_precision": col[3],
                "data_scale": col[4],
                "nullable": col[5]
            })
        
        # Get sample data (last 5 records with TIPODOC = 'DC07')
        cursor.execute("""
            SELECT * FROM (
                SELECT * FROM MANAGER.MNGMCN 
                WHERE MCNTIPODOC = 'DC07'
                ORDER BY MCNNUMEDOC DESC
            ) WHERE ROWNUM <= 5
        """)
        
        # Get column names for the results
        col_names = [desc[0] for desc in cursor.description]
        rows = cursor.fetchall()
        
        sample_data = []
        for row in rows:
            row_dict = {}
            for i, val in enumerate(row):
                # Convert to string for JSON serialization
                if val is not None:
                    row_dict[col_names[i]] = str(val) if not isinstance(val, (int, float)) else val
                else:
                    row_dict[col_names[i]] = None
            sample_data.append(row_dict)
        
        # Get max NUMEDOC for DC07
        cursor.execute("""
            SELECT MAX(MCNNUMEDOC) FROM MANAGER.MNGMCN WHERE MCNTIPODOC = 'DC07'
        """)
        max_numedoc = cursor.fetchone()[0]
        
        return {
            "success": True,
            "tabla": "MANAGER.MNGMCN",
            "total_columnas": len(estructura),
            "max_numedoc_dc07": max_numedoc,
            "estructura": estructura,
            "sample_data_dc07": sample_data,
            "columnas_en_sample": col_names
        }
        
    except Exception as e:
        return {
            "success": False,
            "error": str(e)
        }
    finally:
        if cursor:
            cursor.close()
        if connection:
            connection.close()


@router.get("/mngdoc/estructura")
async def get_mngdoc_estructura():
    """
    Diagnostic endpoint to get the structure of MANAGER.MNGDOC table
    and its relationship with MNGMCN.
    """
    import sys
    sys.path.append('..')
    from oracle_database import get_oracle_connection
    
    connection = None
    cursor = None
    try:
        connection = get_oracle_connection()
        cursor = connection.cursor()
        
        # Get MNGDOC table structure
        cursor.execute("""
            SELECT 
                COLUMN_NAME,
                DATA_TYPE,
                DATA_LENGTH,
                DATA_PRECISION,
                NULLABLE
            FROM ALL_TAB_COLUMNS 
            WHERE OWNER = 'MANAGER' 
            AND TABLE_NAME = 'MNGDOC'
            ORDER BY COLUMN_ID
        """)
        columns = cursor.fetchall()
        
        estructura_mngdoc = []
        for col in columns:
            estructura_mngdoc.append({
                "column_name": col[0],
                "data_type": col[1],
                "data_length": col[2],
                "data_precision": col[3],
                "nullable": col[4]
            })
        
        # Get sample data from MNGDOC for DC07
        cursor.execute("""
            SELECT * FROM (
                SELECT * FROM MANAGER.MNGDOC 
                WHERE DOCTIPO = 'DC07'
                ORDER BY DOCNUMERO DESC
            ) WHERE ROWNUM <= 5
        """)
        
        col_names = [desc[0] for desc in cursor.description]
        rows = cursor.fetchall()
        
        sample_mngdoc = []
        for row in rows:
            row_dict = {}
            for i, val in enumerate(row):
                if val is not None:
                    row_dict[col_names[i]] = str(val) if not isinstance(val, (int, float)) else val
                else:
                    row_dict[col_names[i]] = None
            sample_mngdoc.append(row_dict)
        
        # Check if there's a foreign key or relationship
        cursor.execute("""
            SELECT CONSTRAINT_NAME, CONSTRAINT_TYPE, R_CONSTRAINT_NAME
            FROM ALL_CONSTRAINTS
            WHERE OWNER = 'MANAGER' 
            AND TABLE_NAME IN ('MNGDOC', 'MNGMCN')
            AND CONSTRAINT_TYPE IN ('P', 'R')
        """)
        constraints = cursor.fetchall()
        
        constraints_info = []
        for c in constraints:
            constraints_info.append({
                "constraint_name": c[0],
                "type": c[1],
                "references": c[2]
            })
        
        # Check max DOCNUMERO for DC07
        cursor.execute("""
            SELECT MAX(DOCNUMERO) FROM MANAGER.MNGDOC WHERE DOCTIPO = 'DC07'
        """)
        max_docnumero = cursor.fetchone()[0]
        
        # Compare with MNGMCN max
        cursor.execute("""
            SELECT MAX(MCNNUMEDOC) FROM MANAGER.MNGMCN WHERE MCNTIPODOC = 'DC07'
        """)
        max_mcnnumedoc = cursor.fetchone()[0]
        
        return {
            "success": True,
            "MNGDOC": {
                "descripcion": "Transaccional: Encabezado de Documento",
                "total_columnas": len(estructura_mngdoc),
                "max_docnumero_dc07": max_docnumero,
                "estructura": estructura_mngdoc,
                "sample_data_dc07": sample_mngdoc,
                "columnas": col_names
            },
            "comparacion": {
                "max_MNGDOC_DC07": max_docnumero,
                "max_MNGMCN_DC07": max_mcnnumedoc,
                "nota": "MNGDOC es el encabezado, MNGMCN es el detalle de movimientos"
            },
            "constraints": constraints_info
        }
        
    except Exception as e:
        return {
            "success": False,
            "error": str(e)
        }
    finally:
        if cursor:
            cursor.close()
        if connection:
            connection.close()

