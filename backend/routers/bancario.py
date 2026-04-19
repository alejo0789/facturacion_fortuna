"""
Router bancario (Fase 3 del implementation_plan).

Expone el CRUD de extractos bancarios, la importación de movimientos desde
archivos CSV/Excel (Bancolombia y Davivienda) y el motor de conciliación
automática contra el libro mayor de la cuenta PUC del banco.

Endpoints:
  POST   /bancario/extractos/upload          → sube CSV/XLSX y crea extracto
  GET    /bancario/extractos                 → lista extractos
  GET    /bancario/extractos/{id}            → detalle + transacciones
  DELETE /bancario/extractos/{id}            → borra extracto y sus transacciones
  POST   /bancario/conciliacion/analizar/{extracto_id} → corre scoring
  POST   /bancario/conciliacion/aprobar      → marca transacción como CONCILIADO
  GET    /bancario/reglas                    → lista reglas de conciliación
  POST   /bancario/reglas                    → crea regla
  DELETE /bancario/reglas/{id}               → borra regla

Todas las rutas son multi-tenant — filtran por empresa_id resuelto desde el JWT.
"""
from __future__ import annotations

import csv
import hashlib
import io
import logging
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from typing import List, Optional

from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile
from pydantic import BaseModel
from sqlalchemy import and_, or_, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from core.dependencies import get_current_empresa, get_current_user, require_role
from database import get_db
from models_contabilidad import (
    CuentaBancaria,
    ExtractoBancario,
    LineaAsiento,
    ReglaConciliacion,
    TransaccionBancaria,
    AsientoContable,
)

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/bancario", tags=["bancario"])


# ==========================================================
# Schemas locales
# ==========================================================
class TransaccionBancariaOut(BaseModel):
    id: int
    fecha: date
    descripcion: str
    referencia: Optional[str] = None
    monto: Decimal
    naturaleza: str
    estado_conciliacion: str
    linea_asiento_id: Optional[int] = None

    model_config = {"from_attributes": True}


class ExtractoBancarioOut(BaseModel):
    id: int
    empresa_id: Optional[int] = None
    cuenta_bancaria_id: int
    fecha_inicio: date
    fecha_fin: date
    saldo_inicial: Decimal
    saldo_final: Decimal
    archivo_origen: Optional[str] = None
    estado: str
    total_transacciones: int = 0
    conciliadas: int = 0

    model_config = {"from_attributes": True}


class ExtractoDetalleOut(ExtractoBancarioOut):
    transacciones: List[TransaccionBancariaOut] = []


class UploadExtractoResponse(BaseModel):
    extracto_id: int
    total_transacciones: int
    duplicadas_omitidas: int
    banco: str


class SugerenciaConciliacion(BaseModel):
    transaccion_id: int
    linea_asiento_id: int
    asiento_numero: int
    score: int
    monto_linea: Decimal
    fecha_linea: date
    descripcion_linea: Optional[str] = None
    detalle_match: str


class AnalizarResponse(BaseModel):
    extracto_id: int
    total_transacciones: int
    analizadas: int
    sugerencias: List[SugerenciaConciliacion]
    auto_conciliadas: int


class AprobarConciliacionRequest(BaseModel):
    transaccion_id: int
    linea_asiento_id: int


class ReglaConciliacionOut(BaseModel):
    id: int
    nombre: str
    condicion_descripcion: Optional[str] = None
    condicion_monto_minimo: Optional[Decimal] = None
    condicion_monto_maximo: Optional[Decimal] = None
    cuenta_puc_destino: Optional[str] = None
    crear_asiento_automatico: bool = False
    prioridad: int = 10
    activa: bool

    model_config = {"from_attributes": True}


class ReglaConciliacionCreate(BaseModel):
    nombre: str
    condicion_descripcion: Optional[str] = None
    condicion_monto_minimo: Optional[Decimal] = None
    condicion_monto_maximo: Optional[Decimal] = None
    cuenta_puc_destino: Optional[str] = None
    crear_asiento_automatico: bool = False
    prioridad: int = 10


# ==========================================================
# Parsers CSV / Excel (Bancolombia, Davivienda, Genérico)
# ==========================================================
def _parse_decimal(raw: str) -> Decimal:
    """Convierte strings como '1.234.567,89' o '1,234,567.89' o '-500' a Decimal."""
    if raw is None:
        return Decimal("0")
    s = str(raw).strip().replace(" ", "")
    if not s or s in {"-", "$"}:
        return Decimal("0")
    # remover simbolo moneda
    s = s.replace("$", "").replace("COP", "")
    # detectar formato
    if "," in s and "." in s:
        # 1.234,56 → usa coma decimal
        if s.rfind(",") > s.rfind("."):
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", "")
    elif "," in s:
        # 1234,56 → coma decimal
        s = s.replace(",", ".")
    try:
        return Decimal(s)
    except InvalidOperation:
        return Decimal("0")


def _parse_fecha(raw: str) -> Optional[date]:
    if not raw:
        return None
    s = str(raw).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d", "%d/%m/%y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _hash_transaccion(fecha: date, monto: Decimal, referencia: str, descripcion: str) -> str:
    """Firma estable para evitar duplicados al re-subir el mismo extracto."""
    base = f"{fecha.isoformat()}|{monto}|{(referencia or '').strip()}|{(descripcion or '')[:80].strip()}"
    return hashlib.sha1(base.encode("utf-8")).hexdigest()


def _normalize_header(h: str) -> str:
    return (h or "").strip().lower().replace("  ", " ")


def _detectar_banco(headers: List[str]) -> str:
    lowered = " ".join(_normalize_header(h) for h in headers)
    if "valor consignación" in lowered or "valor consignacion" in lowered:
        return "BANCOLOMBIA"
    if "débito" in lowered and "crédito" in lowered:
        return "DAVIVIENDA"
    if "debito" in lowered and "credito" in lowered:
        return "DAVIVIENDA"
    return "GENERICO"


def _row_to_transaccion(row: dict, banco: str) -> Optional[dict]:
    """Normaliza una fila del CSV al formato interno."""
    keys = {_normalize_header(k): v for k, v in row.items()}

    fecha_raw = (
        keys.get("fecha")
        or keys.get("fecha movimiento")
        or keys.get("fecha transacción")
        or keys.get("fecha transaccion")
    )
    fecha = _parse_fecha(fecha_raw)
    if not fecha:
        return None

    descripcion = (
        keys.get("descripción")
        or keys.get("descripcion")
        or keys.get("detalle")
        or keys.get("concepto")
        or ""
    ).strip()

    referencia = (
        keys.get("referencia")
        or keys.get("documento")
        or keys.get("sucursal")
        or ""
    ).strip() or None

    # Monto: algunos bancos traen débito y crédito en columnas separadas
    debito = _parse_decimal(keys.get("débito") or keys.get("debito") or "0")
    credito = _parse_decimal(keys.get("crédito") or keys.get("credito") or "0")

    if debito or credito:
        if debito > 0:
            monto = debito
            naturaleza = "DEBITO"
        else:
            monto = credito
            naturaleza = "CREDITO"
    else:
        # Columna única "valor" que puede ser negativa
        valor = _parse_decimal(keys.get("valor") or keys.get("monto") or "0")
        if valor == 0:
            return None
        if valor < 0:
            monto = -valor
            naturaleza = "DEBITO"
        else:
            monto = valor
            naturaleza = "CREDITO"

    return {
        "fecha": fecha,
        "descripcion": descripcion or "(sin descripción)",
        "referencia": referencia,
        "monto": monto,
        "naturaleza": naturaleza,
    }


def _leer_csv_bytes(raw: bytes) -> List[dict]:
    """Lee CSV con autodetección de encoding y separador."""
    for encoding in ("utf-8-sig", "utf-8", "latin-1", "cp1252"):
        try:
            texto = raw.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        raise HTTPException(400, "No se pudo decodificar el archivo")

    # Detectar separador: ; o ,
    sniff = texto[:4096]
    delim = ";" if sniff.count(";") > sniff.count(",") else ","

    reader = csv.DictReader(io.StringIO(texto), delimiter=delim)
    return list(reader)


def _leer_xlsx_bytes(raw: bytes) -> List[dict]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise HTTPException(500, "openpyxl no instalado — use CSV o instale openpyxl")

    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(c) if c is not None else "" for c in rows[0]]
    out = []
    for r in rows[1:]:
        if not any(r):
            continue
        out.append({headers[i]: r[i] for i in range(min(len(headers), len(r)))})
    return out


# ==========================================================
# Extractos
# ==========================================================
@router.post("/extractos/upload", response_model=UploadExtractoResponse)
async def upload_extracto(
    cuenta_bancaria_id: int = Form(...),
    archivo: UploadFile = File(...),
    empresa=Depends(get_current_empresa),
    _=Depends(require_role("ADMIN", "CONTADOR", "CONTABILIDAD")),
    db: AsyncSession = Depends(get_db),
):
    """Sube un extracto (CSV/XLSX) y materializa sus transacciones."""
    # Validar cuenta bancaria
    result = await db.execute(
        select(CuentaBancaria).where(
            CuentaBancaria.id == cuenta_bancaria_id,
            CuentaBancaria.empresa_id == empresa.id,
        )
    )
    cuenta = result.scalar_one_or_none()
    if not cuenta:
        raise HTTPException(404, "Cuenta bancaria no encontrada")

    raw = await archivo.read()
    filename = archivo.filename or "extracto"
    lower = filename.lower()

    if lower.endswith(".xlsx") or lower.endswith(".xls"):
        rows = _leer_xlsx_bytes(raw)
    else:
        rows = _leer_csv_bytes(raw)

    if not rows:
        raise HTTPException(400, "El archivo no contiene filas")

    banco = _detectar_banco(list(rows[0].keys()))

    # Normalizar
    normalizadas: List[dict] = []
    for r in rows:
        n = _row_to_transaccion(r, banco)
        if n:
            normalizadas.append(n)

    if not normalizadas:
        raise HTTPException(400, "No se pudo interpretar ninguna fila del extracto")

    fecha_inicio = min(n["fecha"] for n in normalizadas)
    fecha_fin = max(n["fecha"] for n in normalizadas)

    extracto = ExtractoBancario(
        empresa_id=empresa.id,
        cuenta_bancaria_id=cuenta.id,
        fecha_inicio=fecha_inicio,
        fecha_fin=fecha_fin,
        saldo_inicial=Decimal("0"),
        saldo_final=Decimal("0"),
        archivo_origen=filename,
        estado="IMPORTADO",
    )
    db.add(extracto)
    await db.flush()

    # Dedupe: hash(fecha,monto,ref,desc) ya existente en este extracto
    existentes: set[str] = set()
    result = await db.execute(
        select(TransaccionBancaria)
        .join(ExtractoBancario, ExtractoBancario.id == TransaccionBancaria.extracto_id)
        .where(
            ExtractoBancario.empresa_id == empresa.id,
            ExtractoBancario.cuenta_bancaria_id == cuenta.id,
        )
    )
    for t in result.scalars().all():
        existentes.add(_hash_transaccion(t.fecha, t.monto, t.referencia or "", t.descripcion or ""))

    insertadas = 0
    duplicadas = 0
    for n in normalizadas:
        h = _hash_transaccion(n["fecha"], n["monto"], n["referencia"] or "", n["descripcion"])
        if h in existentes:
            duplicadas += 1
            continue
        existentes.add(h)
        db.add(TransaccionBancaria(
            extracto_id=extracto.id,
            fecha=n["fecha"],
            descripcion=n["descripcion"][:500],
            referencia=n["referencia"],
            monto=n["monto"],
            naturaleza=n["naturaleza"],
            estado_conciliacion="NO_CONCILIADO",
        ))
        insertadas += 1

    await db.commit()
    await db.refresh(extracto)
    return UploadExtractoResponse(
        extracto_id=extracto.id,
        total_transacciones=insertadas,
        duplicadas_omitidas=duplicadas,
        banco=banco,
    )


@router.get("/extractos", response_model=List[ExtractoBancarioOut])
async def listar_extractos(
    cuenta_bancaria_id: Optional[int] = None,
    empresa=Depends(get_current_empresa),
    db: AsyncSession = Depends(get_db),
):
    stmt = select(ExtractoBancario).where(ExtractoBancario.empresa_id == empresa.id)
    if cuenta_bancaria_id:
        stmt = stmt.where(ExtractoBancario.cuenta_bancaria_id == cuenta_bancaria_id)
    stmt = stmt.order_by(ExtractoBancario.fecha_fin.desc())
    rows = (await db.execute(stmt.options(selectinload(ExtractoBancario.transacciones)))).scalars().all()

    out = []
    for e in rows:
        total = len(e.transacciones)
        conc = sum(1 for t in e.transacciones if t.estado_conciliacion == "CONCILIADO")
        out.append(ExtractoBancarioOut(
            id=e.id,
            empresa_id=e.empresa_id,
            cuenta_bancaria_id=e.cuenta_bancaria_id,
            fecha_inicio=e.fecha_inicio,
            fecha_fin=e.fecha_fin,
            saldo_inicial=e.saldo_inicial,
            saldo_final=e.saldo_final,
            archivo_origen=e.archivo_origen,
            estado=e.estado,
            total_transacciones=total,
            conciliadas=conc,
        ))
    return out


@router.get("/extractos/{extracto_id}", response_model=ExtractoDetalleOut)
async def detalle_extracto(
    extracto_id: int,
    empresa=Depends(get_current_empresa),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(
        select(ExtractoBancario)
        .where(ExtractoBancario.id == extracto_id, ExtractoBancario.empresa_id == empresa.id)
        .options(selectinload(ExtractoBancario.transacciones))
    )
    e = result.scalar_one_or_none()
    if not e:
        raise HTTPException(404, "Extracto no encontrado")

    total = len(e.transacciones)
    conc = sum(1 for t in e.transacciones if t.estado_conciliacion == "CONCILIADO")
    return ExtractoDetalleOut(
        id=e.id,
        empresa_id=e.empresa_id,
        cuenta_bancaria_id=e.cuenta_bancaria_id,
        fecha_inicio=e.fecha_inicio,
        fecha_fin=e.fecha_fin,
        saldo_inicial=e.saldo_inicial,
        saldo_final=e.saldo_final,
        archivo_origen=e.archivo_origen,
        estado=e.estado,
        total_transacciones=total,
        conciliadas=conc,
        transacciones=[TransaccionBancariaOut.model_validate(t, from_attributes=True) for t in e.transacciones],
    )


@router.delete("/extractos/{extracto_id}")
async def eliminar_extracto(
    extracto_id: int,
    empresa=Depends(get_current_empresa),
    _=Depends(require_role("ADMIN", "CONTADOR")),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(
        select(ExtractoBancario).where(
            ExtractoBancario.id == extracto_id,
            ExtractoBancario.empresa_id == empresa.id,
        )
    )
    e = result.scalar_one_or_none()
    if not e:
        raise HTTPException(404, "Extracto no encontrado")
    await db.delete(e)
    await db.commit()
    return {"ok": True, "extracto_id": extracto_id}


# ==========================================================
# Motor de conciliación (scoring)
# ==========================================================
def _score(transaccion: TransaccionBancaria, linea: LineaAsiento, fecha_linea: date) -> tuple[int, str]:
    """
    Scoring heurístico:
      +50  monto exactamente igual
      +20  monto con tolerancia 1%
      +30  fecha ±3 días
      +15  fecha ±7 días
      +20  referencia contiene NIT o primeras palabras del detalle
    """
    score = 0
    detalles = []

    monto_linea = linea.debito if linea.debito > 0 else linea.credito

    if transaccion.monto == monto_linea:
        score += 50
        detalles.append("monto_exacto")
    elif monto_linea > 0:
        diff_pct = abs(transaccion.monto - monto_linea) / monto_linea
        if diff_pct <= Decimal("0.01"):
            score += 20
            detalles.append("monto_~1%")

    diff_dias = abs((transaccion.fecha - fecha_linea).days)
    if diff_dias <= 3:
        score += 30
        detalles.append(f"fecha_±{diff_dias}d")
    elif diff_dias <= 7:
        score += 15
        detalles.append(f"fecha_±{diff_dias}d")

    ref = (transaccion.referencia or "") + " " + (transaccion.descripcion or "")
    ref_l = ref.lower()
    if linea.nit_tercero and linea.nit_tercero in ref_l:
        score += 20
        detalles.append("nit_match")
    elif linea.detalle:
        primera_palabra = linea.detalle.split()[0].lower() if linea.detalle.split() else ""
        if primera_palabra and len(primera_palabra) >= 4 and primera_palabra in ref_l:
            score += 10
            detalles.append("palabra_match")

    return score, ",".join(detalles)


@router.post("/conciliacion/analizar/{extracto_id}", response_model=AnalizarResponse)
async def analizar_extracto(
    extracto_id: int,
    empresa=Depends(get_current_empresa),
    _=Depends(require_role("ADMIN", "CONTADOR", "CONTABILIDAD")),
    db: AsyncSession = Depends(get_db),
):
    """
    Analiza cada transacción del extracto contra las líneas de asiento
    de la cuenta PUC bancaria. Marca score ≥ 70 como SUGERIDO y score = 100
    como CONCILIADO automático.
    """
    result = await db.execute(
        select(ExtractoBancario)
        .where(ExtractoBancario.id == extracto_id, ExtractoBancario.empresa_id == empresa.id)
        .options(selectinload(ExtractoBancario.transacciones))
    )
    e = result.scalar_one_or_none()
    if not e:
        raise HTTPException(404, "Extracto no encontrado")

    # Cuenta PUC del banco
    result = await db.execute(
        select(CuentaBancaria).where(CuentaBancaria.id == e.cuenta_bancaria_id)
    )
    cuenta = result.scalar_one_or_none()
    if not cuenta:
        raise HTTPException(400, "La cuenta bancaria del extracto no existe")
    cuenta_puc = cuenta.cuenta_puc_codigo

    # Traer líneas candidatas: mismo código PUC, rango de fechas ±10 días, aún no conciliadas
    rango_ini = e.fecha_inicio - timedelta(days=10)
    rango_fin = e.fecha_fin + timedelta(days=10)

    stmt_lineas = (
        select(LineaAsiento, AsientoContable)
        .join(AsientoContable, AsientoContable.id == LineaAsiento.asiento_id)
        .where(
            AsientoContable.empresa_id == empresa.id,
            AsientoContable.estado != "ANULADO",
            LineaAsiento.cuenta_codigo == cuenta_puc,
            AsientoContable.fecha >= rango_ini,
            AsientoContable.fecha <= rango_fin,
        )
    )
    lineas = (await db.execute(stmt_lineas)).all()
    # Quitar las ya conciliadas
    conciliadas_ids = set()
    r2 = await db.execute(
        select(TransaccionBancaria.linea_asiento_id)
        .where(
            TransaccionBancaria.estado_conciliacion == "CONCILIADO",
            TransaccionBancaria.linea_asiento_id.isnot(None),
        )
    )
    for row in r2.all():
        conciliadas_ids.add(row[0])

    candidatas = [(l, a) for (l, a) in lineas if l.id not in conciliadas_ids]

    sugerencias: List[SugerenciaConciliacion] = []
    auto_conciliadas = 0
    analizadas = 0

    # Evitar que dentro de un mismo pase dos transacciones se auto-concilien
    # contra la misma línea (violaría el índice UNIQUE parcial).
    lineas_usadas_en_pase: set[int] = set()

    for t in e.transacciones:
        if t.estado_conciliacion == "CONCILIADO":
            continue
        analizadas += 1

        mejor = None
        mejor_score = 0
        mejor_detalle = ""
        for (linea, asiento) in candidatas:
            if linea.id in lineas_usadas_en_pase:
                continue
            s, det = _score(t, linea, asiento.fecha)
            if s > mejor_score:
                mejor_score = s
                mejor = (linea, asiento)
                mejor_detalle = det

        if not mejor or mejor_score < 70:
            continue

        linea, asiento = mejor
        if mejor_score >= 100:
            t.estado_conciliacion = "CONCILIADO"
            t.linea_asiento_id = linea.id
            auto_conciliadas += 1
            lineas_usadas_en_pase.add(linea.id)
        else:
            t.estado_conciliacion = "SUGERIDO"

        sugerencias.append(SugerenciaConciliacion(
            transaccion_id=t.id,
            linea_asiento_id=linea.id,
            asiento_numero=asiento.numero,
            score=mejor_score,
            monto_linea=(linea.debito if linea.debito > 0 else linea.credito),
            fecha_linea=asiento.fecha,
            descripcion_linea=linea.detalle,
            detalle_match=mejor_detalle,
        ))

    await db.commit()
    return AnalizarResponse(
        extracto_id=e.id,
        total_transacciones=len(e.transacciones),
        analizadas=analizadas,
        sugerencias=sugerencias,
        auto_conciliadas=auto_conciliadas,
    )


@router.post("/conciliacion/aprobar")
async def aprobar_conciliacion(
    data: AprobarConciliacionRequest,
    empresa=Depends(get_current_empresa),
    _=Depends(require_role("ADMIN", "CONTADOR", "CONTABILIDAD")),
    db: AsyncSession = Depends(get_db),
):
    """Confirma manualmente una sugerencia (score ≥ 70) como CONCILIADO."""
    result = await db.execute(
        select(TransaccionBancaria)
        .join(ExtractoBancario, ExtractoBancario.id == TransaccionBancaria.extracto_id)
        .where(
            TransaccionBancaria.id == data.transaccion_id,
            ExtractoBancario.empresa_id == empresa.id,
        )
    )
    t = result.scalar_one_or_none()
    if not t:
        raise HTTPException(404, "Transacción no encontrada")

    # Validar que la línea pertenezca a la misma empresa.
    # FOR UPDATE bloquea la fila hasta el commit → si otra request intenta
    # conciliar la misma `linea_asiento_id` en paralelo, se serializa.
    result = await db.execute(
        select(LineaAsiento)
        .join(AsientoContable, AsientoContable.id == LineaAsiento.asiento_id)
        .where(
            LineaAsiento.id == data.linea_asiento_id,
            AsientoContable.empresa_id == empresa.id,
        )
        .with_for_update()
    )
    linea = result.scalar_one_or_none()
    if not linea:
        raise HTTPException(404, "Línea contable no encontrada")

    # Pre-check defensivo: ¿ya hay una transacción CONCILIADA contra esta línea?
    # El índice UNIQUE parcial `ux_transaccion_bancaria_linea_conciliada` es la
    # garantía real; este chequeo devuelve un 409 legible en lugar de un 500
    # por IntegrityError cuando la segunda request entra tras soltar el lock.
    r_existente = await db.execute(
        select(TransaccionBancaria.id).where(
            TransaccionBancaria.linea_asiento_id == linea.id,
            TransaccionBancaria.estado_conciliacion == "CONCILIADO",
            TransaccionBancaria.id != t.id,
        )
    )
    if r_existente.scalar_one_or_none() is not None:
        raise HTTPException(
            409,
            "La línea contable ya está conciliada con otra transacción bancaria",
        )

    t.estado_conciliacion = "CONCILIADO"
    t.linea_asiento_id = linea.id
    await db.commit()
    return {"ok": True, "transaccion_id": t.id, "linea_asiento_id": linea.id}


@router.post("/conciliacion/rechazar/{transaccion_id}")
async def rechazar_sugerencia(
    transaccion_id: int,
    empresa=Depends(get_current_empresa),
    _=Depends(require_role("ADMIN", "CONTADOR", "CONTABILIDAD")),
    db: AsyncSession = Depends(get_db),
):
    """Devuelve una transacción SUGERIDA al estado NO_CONCILIADO."""
    result = await db.execute(
        select(TransaccionBancaria)
        .join(ExtractoBancario, ExtractoBancario.id == TransaccionBancaria.extracto_id)
        .where(
            TransaccionBancaria.id == transaccion_id,
            ExtractoBancario.empresa_id == empresa.id,
        )
    )
    t = result.scalar_one_or_none()
    if not t:
        raise HTTPException(404, "Transacción no encontrada")

    t.estado_conciliacion = "NO_CONCILIADO"
    t.linea_asiento_id = None
    await db.commit()
    return {"ok": True, "transaccion_id": t.id}


# ==========================================================
# Reglas de conciliación
# ==========================================================
@router.get("/reglas", response_model=List[ReglaConciliacionOut])
async def listar_reglas(
    solo_activas: bool = True,
    empresa=Depends(get_current_empresa),
    db: AsyncSession = Depends(get_db),
):
    stmt = select(ReglaConciliacion).where(ReglaConciliacion.empresa_id == empresa.id)
    if solo_activas:
        stmt = stmt.where(ReglaConciliacion.activa == True)  # noqa: E712
    stmt = stmt.order_by(ReglaConciliacion.prioridad.asc())
    return (await db.execute(stmt)).scalars().all()


@router.post("/reglas", response_model=ReglaConciliacionOut)
async def crear_regla(
    data: ReglaConciliacionCreate,
    empresa=Depends(get_current_empresa),
    _=Depends(require_role("ADMIN", "CONTADOR", "CONTABILIDAD")),
    db: AsyncSession = Depends(get_db),
):
    regla = ReglaConciliacion(
        empresa_id=empresa.id,
        nombre=data.nombre,
        condicion_descripcion=data.condicion_descripcion,
        condicion_monto_minimo=data.condicion_monto_minimo,
        condicion_monto_maximo=data.condicion_monto_maximo,
        cuenta_puc_destino=data.cuenta_puc_destino,
        crear_asiento_automatico=data.crear_asiento_automatico,
        prioridad=data.prioridad,
        activa=True,
    )
    db.add(regla)
    await db.commit()
    await db.refresh(regla)
    return regla


@router.delete("/reglas/{regla_id}")
async def eliminar_regla(
    regla_id: int,
    empresa=Depends(get_current_empresa),
    _=Depends(require_role("ADMIN", "CONTADOR")),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(
        select(ReglaConciliacion).where(
            ReglaConciliacion.id == regla_id,
            ReglaConciliacion.empresa_id == empresa.id,
        )
    )
    regla = result.scalar_one_or_none()
    if not regla:
        raise HTTPException(404, "Regla no encontrada")
    regla.activa = False
    await db.commit()
    return {"ok": True, "regla_id": regla_id}
