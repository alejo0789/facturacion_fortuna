"""
Pagos Router - Module for managing payments and generating consolidado reports.

Strategy for Excel generation:
  - Write data ONLY to the 'info' sheet (row 2+).
  - The 'consolidado' sheet already has formulas like =info!A2, =info!B2...
    so logos, formatting, and column widths are never touched.
"""
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from sqlalchemy.orm import selectinload
from typing import List, Optional
from pydantic import BaseModel
import openpyxl
from io import BytesIO
import threading
from datetime import datetime, date
import os
import re
from urllib.parse import quote

from database import get_db
import models

router = APIRouter()

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

MESES_ES = {
    1: 'ENERO', 2: 'FEBRERO', 3: 'MARZO', 4: 'ABRIL', 5: 'MAYO', 6: 'JUNIO',
    7: 'JULIO', 8: 'AGOSTO', 9: 'SEPTIEMBRE', 10: 'OCTUBRE', 11: 'NOVIEMBRE', 12: 'DICIEMBRE'
}

TEMPLATE_PATH = os.path.join(
    os.path.dirname(__file__), '..', 'Template_consolidado',
    'PROGRAMACION FEBRERO 12 2025.xlsx'
)

# ---------------------------------------------------------------------------
# Schemas
# ---------------------------------------------------------------------------

class ConsolidadoPagosRequest(BaseModel):
    factura_ids: List[int]
    semana_pago: Optional[str] = None   # e.g. "12 DE FEBRERO DE 2025"
    tipo_pago: Optional[str] = "CONSIGNACIÓN"

class FacturaPagoItem(BaseModel):
    factura_id: Optional[int] = None
    documento_contable: str
    valor_pagar: float

class NotaBancariaRequest(BaseModel):
    cuenta_banco: str
    ccosto: str
    destino: str
    detalle: str
    # Campos para una sola factura (compatibilidad)
    factura_id: Optional[int] = None
    documento_contable: Optional[str] = None
    valor_pagar: Optional[float] = None
    # Campo para múltiples facturas
    items: Optional[List[FacturaPagoItem]] = None


# ---------------------------------------------------------------------------
# Helper
# ---------------------------------------------------------------------------

def extract_doc_contable(obs: str) -> str:
    """Extract documento contable code (e.g. DC07-1666) from observaciones text."""
    if not obs:
        return ""
    # Improved regex: looks for DC followed by digits, or DC with hyphen and digits
    m = re.search(r'DC\w*[-\s]?\d+', obs, re.IGNORECASE)
    return m.group(0).upper() if m else ""


# ---------------------------------------------------------------------------
# Endpoint: list facturas EN_TRAMITE
# ---------------------------------------------------------------------------

@router.get("/pagos/facturas-en-tramite")
async def get_facturas_en_tramite(
    db: AsyncSession = Depends(get_db),
    search: Optional[str] = Query(None),
    fecha_desde: Optional[str] = Query(None),
    fecha_hasta: Optional[str] = Query(None),
):
    """
    Return all facturas with estado='EN_TRAMITE' — those sent to Manager for
    payment. Returns enriched data needed for the payments view.
    """
    stmt = (
        select(models.Factura)
        .options(
            selectinload(models.Factura.proveedor),
            selectinload(models.Factura.oficinas_asignadas)
            .selectinload(models.FacturaOficina.oficina),
            selectinload(models.Factura.oficinas_asignadas)
            .selectinload(models.FacturaOficina.contrato),
        )
        .where(models.Factura.estado.in_(['EN_TRAMITE', 'PAGADA']))
        .order_by(models.Factura.status_updated_at.desc())
    )

    if fecha_desde:
        try:
            d = date.fromisoformat(fecha_desde)
            stmt = stmt.where(
                models.Factura.status_updated_at >= datetime.combine(d, datetime.min.time())
            )
        except ValueError:
            pass

    if fecha_hasta:
        try:
            d = date.fromisoformat(fecha_hasta)
            stmt = stmt.where(
                models.Factura.status_updated_at <= datetime.combine(d, datetime.max.time())
            )
        except ValueError:
            pass

    result = await db.execute(stmt)
    facturas = result.scalars().all()

    rows = []
    for f in facturas:
        proveedor_nombre = f.proveedor.nombre if f.proveedor else ''
        proveedor_nit = f.proveedor.nit if f.proveedor else ''

        oficinas_info = []
        for oa in (f.oficinas_asignadas or []):
            oficinas_info.append({
                "oficina_id": oa.oficina_id,
                "oficina_nombre": oa.oficina.nombre if oa.oficina else '',
                "oficina_cod": oa.oficina.cod_oficina if oa.oficina else '',
                "valor": float(oa.valor) if oa.valor else 0,
                "num_contrato": oa.contrato.num_contrato if oa.contrato else None,
                "estado": oa.estado,
                "observaciones": oa.observaciones,
            })

        rows.append({
            "id": f.id,
            "numero_factura": f.numero_factura,
            "fecha_factura": str(f.fecha_factura) if f.fecha_factura else None,
            "fecha_vencimiento": str(f.fecha_vencimiento) if f.fecha_vencimiento else None,
            "valor": float(f.valor) if f.valor else 0,
            "estado": f.estado,
            "status_updated_at": f.status_updated_at.isoformat() if f.status_updated_at else None,
            "observaciones": f.observaciones,
            "url_factura": f.url_factura,
            "proveedor_id": f.proveedor_id,
            "proveedor_nombre": proveedor_nombre,
            "proveedor_nit": proveedor_nit,
            "oficinas": oficinas_info,
            "documento_contable": extract_doc_contable(f.observaciones or ""),
            "cuenta_por_pagar": 23355002,
        })

    # Text search (applied after building rows to search nested fields too)
    if search:
        term = search.lower()
        rows = [
            r for r in rows
            if term in (r.get("numero_factura") or "").lower()
            or term in (r.get("proveedor_nombre") or "").lower()
            or term in (r.get("proveedor_nit") or "").lower()
            or term in (r.get("observaciones") or "").lower()
            or any(term in (o.get("oficina_nombre") or "").lower() for o in r["oficinas"])
        ]

    # --- Check Manager Oracle Approval Status for the list ---
    docs_to_check = []
    for r in rows:
        dc = r.get("documento_contable")
        if dc:
            match = re.search(r'([A-Za-z0-9]+)-(\d+)', dc)
            if match:
                docs_to_check.append((match.group(1).upper(), int(match.group(2))))
                
    aprobados_set = set()
    pagados_set = set()
    
    if docs_to_check:
        try:
            import sys
            sys.path.append('..')
            from oracle_database import get_oracle_connection
            
            conn = get_oracle_connection()
            cursor = conn.cursor()
            
            oracle_details = {}  # {key: MCNDETALLE from account 23355002}
            
            from collections import defaultdict
            grouped = defaultdict(list)
            for t, n in docs_to_check:
                grouped[t].append(n)
                
            for t, nums in grouped.items():
                chunk_size = 900
                for i in range(0, len(nums), chunk_size):
                    chunk = nums[i:i+chunk_size]
                    nums_str = ','.join(map(str, chunk))
                    
                    query = f"""
                        SELECT M.MCNTIPODOC, M.MCNNUMEDOC, M.MCNDIMEORI, 0 as IS_NB01, M.MCNDETALLE
                        FROM MANAGER.MNGMCN M
                        WHERE M.MCNTIPODOC = '{t}' 
                          AND M.MCNNUMEDOC IN ({nums_str})
                          AND TRIM(M.MCNCUENTA) = '23355002'
                        UNION ALL
                        SELECT M.MCNTIPCRU2, M.MCNNUMCRU2, 0 as MCNDIMEORI, 1 as IS_NB01, '' as MCNDETALLE
                        FROM MANAGER.MNGMCN M
                        WHERE M.MCNTIPODOC = 'NB01' 
                          AND M.MCNTIPCRU2 = '{t}' 
                          AND M.MCNNUMCRU2 IN ({nums_str})
                    """
                    cursor.execute(query)
                    for row_or in cursor.fetchall():
                        tipo_o, num_o, dimeori_o, is_nb01, mcn_detalle = row_or
                        key = f"{str(tipo_o).strip()}-{int(num_o)}"
                        
                        if is_nb01 == 1:
                            pagados_set.add(key)
                        else:
                            if mcn_detalle:
                                oracle_details[key] = str(mcn_detalle).strip()
                            if dimeori_o is not None and float(dimeori_o) > 0:
                                aprobados_set.add(key)

            cursor.close()
            conn.close()
        except Exception as e:
            print(f"Error checking oracle approvals in batch: {e}")
            
    # Assign the state and updated observations back to rows
    for r in rows:
        dc = r.get("documento_contable")
        es_aprobado = False
        es_pagada = False
        if dc:
            match = re.search(r'([A-Za-z0-9]+)-(\d+)', dc)
            if match:
                key = f"{match.group(1).upper()}-{int(match.group(2))}"
                es_aprobado = key in aprobados_set
                es_pagada = key in pagados_set
                
                # IMPROVED: ONLY overwrite if current observation doesn't already have the DC reference
                # This preserves your edited details (e.g. "Factura Claro...")
                current_obs = r.get("observaciones") or ""
                if key in oracle_details and "Ref Doc:" not in current_obs:
                    r["observaciones"] = oracle_details[key]
                
        r["es_aprobado_manager"] = es_aprobado
        r["es_pagada_manager"] = es_pagada

    return rows


# ---------------------------------------------------------------------------
# Endpoint: generate Programación de Pagos Excel
# ---------------------------------------------------------------------------

@router.post("/pagos/consolidado-programacion")
async def generar_consolidado_programacion(
    request: ConsolidadoPagosRequest,
    db: AsyncSession = Depends(get_db),
):
    """
    Generate the 'Programación de Pagos' Excel for the selected facturas.

    KEY: writes data ONLY to the 'info' sheet (rows 2+).
    The 'consolidado' sheet pulls data via formulas (=info!A2, =info!B2, ...)
    so logos, column widths, and all formatting remain intact.

    info sheet layout (row 1 = headers already present in template):
      A: ITEM
      B: VALOR NETO A PAGAR
      C: No. FACTURA Y/O CUENTA DE COBRO
      D: CUENTA POR PAGAR  (always 23355002)
      E: CC/NIT
      F: BENEFICIARIO
      G: BANCO             (left blank — user fills)
      H: TIPO DE CUENTA    (left blank — user fills)
      I: CUENTA            (left blank — user fills)
      J: OBSERVACIÓN
      K: DOCUMENTO CONTABLE
    """
    if not request.factura_ids:
        raise HTTPException(status_code=400, detail="No se seleccionaron facturas")

    # Fetch facturas
    stmt = (
        select(models.Factura)
        .options(
            selectinload(models.Factura.proveedor),
            selectinload(models.Factura.oficinas_asignadas)
            .selectinload(models.FacturaOficina.oficina),
            selectinload(models.Factura.oficinas_asignadas)
            .selectinload(models.FacturaOficina.contrato),
        )
        .where(models.Factura.id.in_(request.factura_ids))
        .order_by(models.Factura.id)
    )
    result = await db.execute(stmt)
    facturas = result.scalars().all()

    if not facturas:
        raise HTTPException(status_code=404, detail="No se encontraron las facturas seleccionadas")

    # Load the template — keep everything as-is
    try:
        wb = openpyxl.load_workbook(TEMPLATE_PATH)
    except FileNotFoundError:
        raise HTTPException(
            status_code=500,
            detail=f"Template no encontrado: {TEMPLATE_PATH}"
        )

    if 'info' not in wb.sheetnames:
        raise HTTPException(
            status_code=500,
            detail="El template no contiene la hoja 'info'. Verifica el archivo."
        )

    ws_info = wb['info']

    # --- Configurar conexión a Oracle para traer el detalle exacto de Manager ---
    import sys
    sys.path.append('..')
    from oracle_database import get_oracle_connection
    
    oracle_conn = None
    oracle_cursor = None
    try:
        oracle_conn = get_oracle_connection()
        oracle_cursor = oracle_conn.cursor()
    except Exception as e:
        print(f"⚠ Warning: No se pudo conectar a Oracle para leer los detalles desde Manager: {e}")

    # ── Clear old data rows in 'info' (row 1 = headers, keep untouched) ──
    for row_idx in range(2, ws_info.max_row + 1):
        for col_idx in range(1, 12):      # columns A–K
            ws_info.cell(row=row_idx, column=col_idx).value = None

    # ── Write data rows: info row 2 = item 1, info row 3 = item 2, … ──
    for item_num, factura in enumerate(facturas, start=1):
        info_row = item_num + 1   # data starts at row 2 of info sheet

        proveedor_nit    = (factura.proveedor.nit    or "").strip() if factura.proveedor else ""
        proveedor_nombre = (factura.proveedor.nombre or "").strip() if factura.proveedor else ""
        numero_factura   = factura.numero_factura or ""
        valor            = float(factura.valor) if factura.valor else 0
        observaciones    = factura.observaciones or ""
        doc_contable     = extract_doc_contable(observaciones)

        # Buscar el detalle exacto en Manager si tenemos documento contable
        observacion_excel = observaciones
        if doc_contable and oracle_cursor:
            match = re.search(r'([A-Za-z0-9]+)-(\d+)', doc_contable)
            if match:
                tipo_doc = match.group(1).upper()
                try:
                    num_doc = int(match.group(2))
                    oracle_cursor.execute(
                        "SELECT DOCDETALLE FROM MANAGER.MNGDOC WHERE DOCTIPO = :tipo AND DOCNUMERO = :numero",
                        {'tipo': tipo_doc, 'numero': num_doc}
                    )
                    mngdoc_row = oracle_cursor.fetchone()
                    if mngdoc_row and mngdoc_row[0]:
                        observacion_excel = str(mngdoc_row[0]).strip()
                except Exception as e:
                    print(f"Error consultando detalle en Manager para {doc_contable}: {e}")

        ws_info.cell(row=info_row, column=1).value  = item_num          # A: ITEM
        ws_info.cell(row=info_row, column=2).value  = valor             # B: VALOR NETO A PAGAR
        ws_info.cell(row=info_row, column=3).value  = numero_factura    # C: No. FACTURA
        ws_info.cell(row=info_row, column=4).value  = 23355002          # D: CUENTA POR PAGAR
        ws_info.cell(row=info_row, column=5).value  = proveedor_nit     # E: CC/NIT
        ws_info.cell(row=info_row, column=6).value  = proveedor_nombre  # F: BENEFICIARIO
        # G (col 7) BANCO          → blank, user fills
        # H (col 8) TIPO DE CUENTA → blank, user fills
        # I (col 9) CUENTA         → blank, user fills
        ws_info.cell(row=info_row, column=10).value = observacion_excel  # J: OBSERVACIÓN (DE MANAGER)
        ws_info.cell(row=info_row, column=11).value = doc_contable       # K: DOCUMENTO CONTABLE

    # ── Update header text in 'consolidado' (safe: only text cells, no images) ──
    if 'consolidado' in wb.sheetnames:
        ws_cons = wb['consolidado']
        now = datetime.now()
        semana_pago = request.semana_pago or f"{now.day} DE {MESES_ES[now.month]} DE {now.year}"
        ws_cons['C5'] = f"PROGRAMACIÓN DE PAGOS: DE {semana_pago}"
        if request.tipo_pago:
            ws_cons['G5'] = request.tipo_pago

    # ── Generate filename and return ──
    now = datetime.now()
    dia = str(now.day).zfill(2)
    mes = MESES_ES[now.month]
    anio = now.year
    filename = f"PROGRAMACION-PAGOS-{dia}-{mes}-{anio}.xlsx"

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    encoded_filename = quote(filename)
    
    # Cerrar conexion de oracle si se abrio
    if oracle_cursor:
        try: oracle_cursor.close()
        except: pass
    if oracle_conn:
        try: oracle_conn.close()
        except: pass

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}",
            "Access-Control-Expose-Headers": "Content-Disposition",
        },
    )


# ---------------------------------------------------------------------------
# Endpoint: Verify Manager Causation (Oracle query)
# ---------------------------------------------------------------------------

@router.get("/pagos/causacion-manager/{documento_contable}")
def get_manager_causation_details(documento_contable: str):
    """
    Fetch the ACTUAL causation records directly from Manager's Oracle DB 
    using the Documento Contable (e.g. 'DC07-1666').
    Select ONLY. NO INSERTS.
    """
    import sys
    sys.path.append('..')
    from oracle_database import get_oracle_connection
    
    # Parse document contable (e.g. DC07-1666 -> tipo='DC07', numero=1666)
    match = re.search(r'([A-Za-z0-9]+)-(\d+)', documento_contable)
    if not match:
        raise HTTPException(
            status_code=400, 
            detail=f"Formato de documento contable inválido: '{documento_contable}'. Esperado ej. DC07-1666"
        )
        
    tipo_doc = match.group(1).upper()
    try:
        numero_doc = int(match.group(2))
    except ValueError:
        raise HTTPException(status_code=400, detail="Número de documento inválido")
        
    connection = None
    cursor = None
    try:
        connection = get_oracle_connection()
        cursor = connection.cursor()
        
        # MNGDOC = header (we can check if it exists)
        # MNGMCN = details (this is what the user wants: Cuenta, Tipo, C.Costo, Destino, Valor, Detalle)
        
        query = """
            WITH related_docs AS (
                SELECT MCNTIPODOC, MCNNUMEDOC
                FROM MANAGER.MNGMCN
                WHERE MCNTIPODOC = :tipo AND MCNNUMEDOC = :numero
                UNION
                SELECT MCNTIPODOC, MCNNUMEDOC
                FROM MANAGER.MNGMCN
                WHERE MCNTIPCRU2 = :tipo AND MCNNUMCRU2 = :numero
            )
            SELECT 
                M.MCNCUENTA, 
                M.MCNTIPODOC, 
                M.MCNCCOSTO, 
                M.MCNDESTINO, 
                M.MCNVALDEBI, 
                M.MCNVALCRED, 
                M.MCNDETALLE,
                M.MCNINDINV,
                M.MCNDIMEORI,
                M.MCNNUMEDOC
            FROM MANAGER.MNGMCN M
            JOIN related_docs R ON M.MCNTIPODOC = R.MCNTIPODOC AND M.MCNNUMEDOC = R.MCNNUMEDOC
            ORDER BY CASE WHEN M.MCNTIPODOC = :tipo THEN 1 ELSE 2 END, M.MCNTIPODOC, M.MCNNUMEDOC, M.MCNREG
        """
        
        cursor.execute(query, {'tipo': tipo_doc, 'numero': numero_doc})
        rows = cursor.fetchall()
        
        if not rows:
            return {
                "success": False,
                "message": f"No se encontró causación en Manager para {documento_contable}",
                "data": []
            }
            
        # Format the result
        detalles = []
        es_aprobado = False
        saldo_cxp = 0.0

        for r in rows:
            cuenta, tipo, ccosto, destino, debito, credito, detalle, indinv, dimeori, num_doc = r
            
            # Revisar si para la 23355002 ya está aprobado y acumular saldo
            if str(cuenta).strip().startswith('23355002'):
                if dimeori is not None and float(dimeori) > 0:
                    es_aprobado = True
                
                # Para pasivos, credito suma la deuda, debito la resta.
                v_cred = float(credito) if credito else 0.0
                v_deb  = float(debito) if debito else 0.0
                saldo_cxp += (v_cred - v_deb)

            # Determine if we show debito or credito as the valor
            valor = debito if debito and float(debito) > 0 else credito
            tipo_movimiento = "DEBITO" if debito and float(debito) > 0 else "CREDITO"
            
            # Combine tipo and num
            t_str = str(tipo).strip()
            n_val = int(num_doc) if num_doc else 0
            
            detalles.append({
                "cuenta": str(cuenta),
                "tipo_doc": f"{t_str}-{n_val}",
                "tipo": tipo_movimiento,
                "ccosto": str(ccosto) if ccosto else "",
                "destino": str(destino) if destino else "",
                "valor": float(valor) if valor else 0,
                "detalle": str(detalle) if detalle else ""
            })
            
        return {
            "success": True,
            "message": f"Detalles cargados desde Manager",
            "documento": documento_contable,
            "es_aprobado": es_aprobado,
            "saldo_pendiente": saldo_cxp,
            "pagado": saldo_cxp <= 0.01 and len(detalles) > 0, 
            "data": detalles
        }
        
    except Exception as e:
        print(f"Error querying Manager: {e}")
        raise HTTPException(status_code=500, detail=f"Error consultando Manager: {str(e)}")
    finally:
        if cursor:
            try:
                cursor.close()
            except: pass
        if connection:
            try:
                connection.close()
            except: pass


# ---------------------------------------------------------------------------
# Endpoint: Approve Manager Causation (Oracle update)
# ---------------------------------------------------------------------------

@router.put("/pagos/causacion-manager/{documento_contable}/aprobar")
def aprobar_manager_causation(documento_contable: str):
    """
    Aprobar facturas pendientes en Oracle DB directamente
    modificando la cuenta 23355002.
    """
    import sys
    sys.path.append('..')
    from oracle_database import get_oracle_connection
    import datetime
    
    match = re.search(r'([A-Za-z0-9]+)-(\d+)', documento_contable)
    if not match:
        raise HTTPException(status_code=400, detail="Formato inválido")
        
    tipo_doc = match.group(1).upper()
    try:
        numero_doc = int(match.group(2))
    except ValueError:
        raise HTTPException(status_code=400, detail="Número de documento inválido")
        
    connection = None
    cursor = None
    try:
        connection = get_oracle_connection()
        cursor = connection.cursor()
        
        # 1. Obtener saldo credito
        query_sel = """
            SELECT MCNSALDOCR FROM MANAGER.MNGMCN 
            WHERE MCNTIPODOC = :tipo 
              AND MCNNUMEDOC = :numero 
              AND TRIM(MCNCUENTA) = '23355002'
        """
        cursor.execute(query_sel, {'tipo': tipo_doc, 'numero': numero_doc})
        row = cursor.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Cuenta x pagar 23355002 no encontrada en el documento.")
            
        saldo_cr = float(row[0] or 0)
        
        # 2. Hacer update
        query_upd = """
            UPDATE MANAGER.MNGMCN
            SET MCNDIMEORI = :saldo,
                MCNINDINV = '.',
                MCNMODUSER = 'WEBAPP',
                MCNMODFEC = SYSDATE
            WHERE MCNTIPODOC = :tipo 
              AND MCNNUMEDOC = :numero 
              AND TRIM(MCNCUENTA) = '23355002'
        """
        cursor.execute(query_upd, {'saldo': saldo_cr, 'tipo': tipo_doc, 'numero': numero_doc})
        connection.commit()
        
        return {"success": True, "message": "Documento aprobado exitosamente en Manager."}
        
    except Exception as e:
        if connection:
            try: connection.rollback()
            except: pass
        print(f"Error approving Manager: {e}")
        raise HTTPException(status_code=500, detail=f"Error al aprobar en Manager: {str(e)}")
    finally:
        if cursor:
            try: cursor.close()
            except: pass
        if connection:
            try: connection.close()
            except: pass


# ---------------------------------------------------------------------------
# Endpoints: Notas Bancarias
# ---------------------------------------------------------------------------

@router.get("/pagos/parametros-nota-bancaria")
def get_parametros_nota_bancaria():
    import sys
    sys.path.append('..')
    from oracle_database import get_oracle_connection
    
    conn = None
    cursor = None
    try:
        conn = get_oracle_connection()
        cursor = conn.cursor()
        
        cursor.execute("SELECT CTACODIGO, CTANOMBRE FROM MANAGER.MNGCTA")
        cuentas_raw = cursor.fetchall()
        cuentas = [{"codigo": c[0].strip(), "nombre": c[1].strip() if c[1] else ""} for c in cuentas_raw]
        
        cursor.execute("SELECT CCOCODIGO, CCONOMBRE FROM MANAGER.MNGCCO")
        ccostos_raw = cursor.fetchall()
        ccostos = [{"codigo": c[0].strip(), "nombre": c[1].strip() if c[1] else ""} for c in ccostos_raw]
        
        destinos = []
        try:
            cursor.execute("SELECT DSTCODIGO, DSTNOMBRE FROM MANAGER.MNGDST")
            destinos_raw = cursor.fetchall()
            destinos = [{"codigo": c[0].strip(), "nombre": c[1].strip() if c[1] else ""} for c in destinos_raw]
        except:
            # Table might not exist in local docker
            destinos = [{"codigo": "001", "nombre": "Sede Administrativa"}]

        return {
            "success": True,
            "cuentas": cuentas,
            "ccostos": ccostos,
            "destinos": destinos
        }
    except Exception as e:
        print(f"Error parametros: {e}")
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if cursor:
            try: cursor.close()
            except: pass
        if conn:
            try: conn.close()
            except: pass


@router.post("/pagos/crear-nota-bancaria")
async def crear_nota_bancaria(req: NotaBancariaRequest, db: AsyncSession = Depends(get_db)):
    import sys
    sys.path.append('..')
    from oracle_database import get_oracle_connection
    import datetime
    
    # 1. Normalizar items (si viene uno solo, lo convertimos a lista para procesar igual)
    items_pago = []
    if req.items:
        items_pago = req.items
    elif req.documento_contable and req.valor_pagar:
        items_pago = [FacturaPagoItem(
            factura_id=req.factura_id,
            documento_contable=req.documento_contable,
            valor_pagar=req.valor_pagar
        )]
        
    if not items_pago:
        raise HTTPException(status_code=400, detail="No se proporcionaron facturas para pagar")

    conn = None
    cursor = None
    try:
        conn = get_oracle_connection()
        cursor = conn.cursor()
        
        # 2. Obtener el VINCULADO del primer item para la cabecera del documento
        first_item = items_pago[0]
        match_first = re.search(r'([A-Za-z0-9]+)-(\d+)', first_item.documento_contable)
        if not match_first:
            raise HTTPException(status_code=400, detail=f"Formato inválido: {first_item.documento_contable}")
            
        tipo_primero = match_first.group(1).upper()
        num_primero = int(match_first.group(2))
        
        cursor.execute("SELECT DOCVINCULA FROM MANAGER.MNGDOC WHERE DOCTIPO = :t AND DOCNUMERO = :n", 
                      {'t': tipo_primero, 'n': num_primero})
        row_vinc = cursor.fetchone()
        nit_cabecera = row_vinc[0].strip() if row_vinc and row_vinc[0] else '.'
        
        # 3. Obtener nuevo consecutivo NB01
        cursor.execute("SELECT NVL(MAX(DOCNUMERO), 0) + 1 FROM MANAGER.MNGDOC WHERE DOCTIPO = 'NB01'")
        nb_num = int(cursor.fetchone()[0])
        
        # 4. Insertar Cabecera MNGDOC
        cursor.execute('''
            INSERT INTO MANAGER.MNGDOC (
                DOCTIPO, DOCNUMERO, DOCFECHA, DOCNEWUSER, DOCESTADO, DOCVINCULA, DOCDETALLE
            ) VALUES (
                'NB01', :1, SYSDATE, 'WEBAPP  ', 'a', :2, :3
            )
        ''', (nb_num, nit_cabecera.ljust(15), req.detalle[:100]))
        
        total_pagar = 0.0
        registro_actual = 1
        facturas_ids_procesados = []

        # 5. Insertar MOVIMIENTOS (MNGMCN) - Par Débito/Crédito por cada factura
        for item in items_pago:
            match = re.search(r'([A-Za-z0-9]+)-(\d+)', item.documento_contable)
            if not match: continue
            
            tipo_fac = match.group(1).upper()
            num_fac = int(match.group(2))
            
            # Consultar Ccosto y Destino original de esta factura
            cursor.execute('''
                SELECT M.MCNCCOSTO, M.MCNDESTINO, D.DOCVINCULA
                FROM MANAGER.MNGMCN M
                JOIN MANAGER.MNGDOC D ON M.MCNTIPODOC = D.DOCTIPO AND M.MCNNUMEDOC = D.DOCNUMERO
                WHERE M.MCNTIPODOC = :tfac AND M.MCNNUMEDOC = :nfac AND M.MCNCUENTA LIKE '2335%' AND ROWNUM = 1
            ''', {'tfac': tipo_fac, 'nfac': num_fac})
            c_info = cursor.fetchone()
            
            cc_orig = c_info[0].strip() if c_info and c_info[0] else '.'
            ds_orig = c_info[1].strip() if c_info and c_info[1] else '.'
            nit_fac = c_info[2].strip() if c_info and c_info[2] else nit_cabecera
            
            # A. Insert Detalle Débito (Referenciando cada factura)
            cursor.execute('''
                INSERT INTO MANAGER.MNGMCN (
                    MCNTIPODOC, MCNNUMEDOC, MCNREG, 
                    MCNCUENTA, MCNVALDEBI, MCNVALCRED,
                    MCNDETALLE, MCNINDINV, MCNDIMEORI, MCNCCOSTO, MCNDESTINO,
                    MCNCLACRU1, MCNTIPCRU1, MCNNUMCRU1, MCNCUOCRU1,
                    MCNCLACRU2, MCNTIPCRU2, MCNNUMCRU2, MCNCUOCRU2,
                    MCNEMPRESA, MCNESTADO, MCNFECINI, MCNPLAZO, MCNREF1, MCNREF2, MCNVINCULA
                ) VALUES (
                    'NB01', :num, :reg, 
                    '23355002      ', :val, 0,
                    :det, 'E', 0, :cc, :dst,
                    '    ', 'NB01', :num, 0,
                    '0000', :tfac, :nfac, 0,
                    '101', 'a', SYSDATE, 0, '.', '.', :nit
                )
            ''', {
                'num': nb_num, 'reg': registro_actual, 'val': item.valor_pagar, 'det': req.detalle[:100], 
                'cc': cc_orig.ljust(10), 'dst': ds_orig.ljust(10),
                'tfac': tipo_fac.ljust(4), 'nfac': num_fac, 'nit': nit_fac.ljust(15)
            })
            
            registro_actual += 1
            
            # B. Insert Detalle Crédito (Salida de Banco para ESTA factura)
            cursor.execute('''
                INSERT INTO MANAGER.MNGMCN (
                    MCNTIPODOC, MCNNUMEDOC, MCNREG, 
                    MCNCUENTA, MCNVALDEBI, MCNVALCRED,
                    MCNDETALLE, MCNINDINV, MCNDIMEORI, MCNCCOSTO, MCNDESTINO,
                    MCNCLACRU1, MCNTIPCRU1, MCNNUMCRU1, MCNCUOCRU1,
                    MCNCLACRU2, MCNTIPCRU2, MCNNUMCRU2, MCNCUOCRU2,
                    MCNEMPRESA, MCNESTADO, MCNFECINI, MCNPLAZO, MCNREF1, MCNREF2, MCNVINCULA
                ) VALUES (
                    'NB01', :num, :reg, 
                    :cta, 0, :val,
                    :det, 'E', 0, :cc, :dst,
                    '    ', '    ', 0, 0,
                    '    ', '    ', 0, 0,
                    '101', 'a', SYSDATE, 0, '.', '.', :nit
                )
            ''', {
                'num': nb_num, 'reg': registro_actual, 'cta': req.cuenta_banco.ljust(14), 'val': item.valor_pagar, 
                'det': req.detalle[:100], 'cc': req.ccosto.ljust(10), 'dst': req.destino.ljust(10), 
                'nit': nit_fac.ljust(15)
            })
            
            total_pagar += item.valor_pagar
            registro_actual += 1
            if item.factura_id:
                facturas_ids_procesados.append(item.factura_id)

        conn.commit()
        
        # 6. Actualizar estados en DB local
        if facturas_ids_procesados:
            from sqlalchemy import update
            import models
            await db.execute(
                update(models.Factura)
                .where(models.Factura.id.in_(facturas_ids_procesados))
                .values(estado='PAGADA', status_updated_at=datetime.datetime.utcnow())
            )
            await db.commit()
            
        return {
            "success": True, 
            "message": f"Nota Bancaria NB01-{nb_num} creada exitosamente con {len(items_pago)} facturas y sus respectivas salidas de banco.",
            "nb_numero": f"NB01-{nb_num}",
            "valor_total": total_pagar
        }
    except Exception as e:
        if conn: conn.rollback()
        raise e
    finally:
        if cursor: cursor.close()
        if conn: conn.close()


@router.get("/pagos/nota-bancaria/{tipo}/{numero}/pdf")
def get_nota_bancaria_pdf(tipo: str, numero: int):
    """
    Genera el PDF de una Nota Bancaria, con descripcion de cuenta y vinculado por fila.
    """
    import sys
    sys.path.append('..')
    from oracle_database import get_oracle_connection
    from utils.pdf_generator import generar_pdf_nb01
    from fastapi.responses import Response

    conn = None
    try:
        conn = get_oracle_connection()
        cursor = conn.cursor()

        # ── 1. Movimientos con nombre de cuenta y nombre de tercero por fila ──
        try:
            cursor.execute('''
                SELECT
                    M.MCNCUENTA,
                    M.MCNVALDEBI,
                    M.MCNVALCRED,
                    M.MCNDETALLE,
                    M.MCNCCOSTO,
                    M.MCNDESTINO,
                    M.MCNTIPCRU2,
                    M.MCNNUMCRU2,
                    NVL(M.MCNVINCULA, D.DOCVINCULA)  AS vinculado_nit,
                    NVL(C.CLINOMBRE, '.')             AS nombre_tercero,
                    D.DOCFECHA,
                    NVL(CT.CTANOMBRE, M.MCNCUENTA)   AS nombre_cuenta,
                    DFAC.DOCDETALLE                  AS detalle_factura
                FROM MANAGER.MNGMCN M
                JOIN MANAGER.MNGDOC D  ON M.MCNTIPODOC = D.DOCTIPO AND M.MCNNUMEDOC = D.DOCNUMERO
                LEFT JOIN MANAGER.MNGDOC DFAC ON M.MCNTIPCRU2 = DFAC.DOCTIPO AND M.MCNNUMCRU2 = DFAC.DOCNUMERO
                LEFT JOIN MANAGER.MNGCLI C  ON NVL(M.MCNVINCULA, D.DOCVINCULA) = C.CLINIT
                LEFT JOIN MANAGER.MNGCTA CT ON TRIM(M.MCNCUENTA) = TRIM(CT.CTACODIGO)
                WHERE M.MCNTIPODOC = :tipo AND M.MCNNUMEDOC = :num
                ORDER BY M.MCNREG
            ''', {'tipo': tipo.upper(), 'num': numero})
        except Exception as e:
            if 'ORA-00942' in str(e):
                # Fallback local Docker (probablemente falte MNGCLI, pero MNGCTA y MNGDOC sí existen)
                cursor.execute('''
                    SELECT
                        M.MCNCUENTA,
                        M.MCNVALDEBI,
                        M.MCNVALCRED,
                        M.MCNDETALLE,
                        M.MCNCCOSTO,
                        M.MCNDESTINO,
                        M.MCNTIPCRU2,
                        M.MCNNUMCRU2,
                        NVL(M.MCNVINCULA, D.DOCVINCULA)  AS vinculado_nit,
                        'SIN NOMBRE (LOCAL)'              AS nombre_tercero,
                        D.DOCFECHA,
                        NVL(CT.CTANOMBRE, M.MCNCUENTA)   AS nombre_cuenta,
                        DFAC.DOCDETALLE                  AS detalle_factura
                    FROM MANAGER.MNGMCN M
                    JOIN MANAGER.MNGDOC D  ON M.MCNTIPODOC = D.DOCTIPO AND M.MCNNUMEDOC = D.DOCNUMERO
                    LEFT JOIN MANAGER.MNGDOC DFAC ON M.MCNTIPCRU2 = DFAC.DOCTIPO AND M.MCNNUMCRU2 = DFAC.DOCNUMERO
                    LEFT JOIN MANAGER.MNGCTA CT ON TRIM(M.MCNCUENTA) = TRIM(CT.CTACODIGO)
                    WHERE M.MCNTIPODOC = :tipo AND M.MCNNUMEDOC = :num
                    ORDER BY M.MCNREG
                ''', {'tipo': tipo.upper(), 'num': numero})
            else:
                raise e

        rows = cursor.fetchall()
        if not rows:
            raise HTTPException(status_code=404, detail="Soporte contable no encontrado en Manager")

        # ── 2. Cabecera: usar el PRIMER débito (proveedor, no banco) ──
        primera_fila_debito = next((r for r in rows if (r[1] or 0) > 0), rows[0])
        nit_cabecera    = str(primera_fila_debito[8]).strip() if primera_fila_debito[8] else '.'
        nombre_cabecera = str(primera_fila_debito[9]).strip() if primera_fila_debito[9] else 'PROVEEDOR'
        fecha_cabecera  = rows[0][10].strftime("%Y/%m/%d") if rows[0][10] else 'N/A'

        nb_data = {
            'numero_doc':       str(numero),
            'fecha':            fecha_cabecera,
            'detalle_cabecera': str(rows[0][3]).strip() if rows[0][3] else '.',
            'nit':              nit_cabecera,
            'nombre_tercero':   nombre_cabecera,
            'ccosto_cabecera':  str(rows[0][4]).strip() if rows[0][4] else '.',
            'destino_cabecera': str(rows[0][5]).strip() if rows[0][5] else '.',
            'valor_total':      sum(float(r[1]) for r in rows if r[1]),
            'detalles': []
        }
        
        import re

        for r in rows:
            debito  = float(r[1]) if r[1] else 0.0
            credito = float(r[2]) if r[2] else 0.0
            tcru2   = str(r[6]).strip() if r[6] else ''
            ncru2   = str(int(r[7])) if r[7] else '0'
            doc_cruce = f"{tcru2}-{ncru2}" if tcru2 else '.'

            nit_fila    = str(r[8]).strip() if r[8] else '.'
            
            detalle_original = str(r[3]).strip() if r[3] else ''
            nombre_cta_bd    = str(r[11]).strip() if r[11] else ''
            detalle_factura  = str(r[12]).strip() if len(r) > 12 and r[12] else ''
            cuenta_num       = str(r[0]).strip()
            
            # Lógica para la descripción:
            # - Si es Débito (CxP / Factura): Detalles originales de LA FACTURA (MNGDOC DFAC) truncados a 30 caracteres
            # - Si es Crédito (Banco): Nombre de la cuenta, cortado en 'CT'
            if credito > 0:
                desc = nombre_cta_bd if nombre_cta_bd and nombre_cta_bd != cuenta_num else "CUENTA BANCO"
                # Limpiar la palabra despues de CT (ej: "BANCO BBVA CT 001305..." -> "BANCO BBVA CT")
                match_ct = re.search(r'(?i)(.*?\bCT\b)', desc)
                if match_ct:
                    desc = match_ct.group(1).upper()
            else:
                desc = detalle_factura[:30] if detalle_factura else (nombre_cta_bd[:30] if nombre_cta_bd else detalle_original[:30])

            nb_data['detalles'].append({
                'cuenta':    cuenta_num,
                'desc':      desc,
                'nit':       nit_fila,
                'doc_cruce': doc_cruce,
                'debito':    debito,
                'credito':   credito,
                'ccosto':    str(r[4]).strip() if r[4] else '.',
                'destino':   str(r[5]).strip() if r[5] else '.',
            })

        pdf_bytes = generar_pdf_nb01(nb_data)
        return Response(content=pdf_bytes, media_type="application/pdf")

    except Exception as e:
        print("Error generando PDF Nota Bancaria:", e)
        raise HTTPException(status_code=500, detail=f"Error generando PDF: {str(e)}")
