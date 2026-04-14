import openpyxl
wb = openpyxl.load_workbook('backend/Template_consolidado/template_relacion_facturas.xlsx', data_only=True)
ws = wb['info']
print(f"L1: {ws['L1'].value}")
print(f"M1: {ws['M1'].value}")
print(f"U1: {ws['U1'].value}")
